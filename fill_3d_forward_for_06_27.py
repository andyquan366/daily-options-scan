import pandas as pd
from datetime import datetime, timedelta
import yfinance as yf
from openpyxl import load_workbook
import re

file_name = "option_activity_log.xlsx"
wb = load_workbook(file_name)

# 指定需要补齐的日期
target_date = datetime(2025, 6, 27).date()
base_day = target_date + timedelta(days=3)  # 目标基准日，即30号

pattern = re.compile(r"^\d{4}-\d{2}$")  # 匹配 yyyy-mm 格式

def get_price_realtime(ticker, target_date, max_lookback=14):
    from datetime import timedelta
    for i in range(max_lookback):
        try_date = target_date - timedelta(days=i)
        if try_date.weekday() >= 5:  # 跳过周末
            continue
        try:
            stock = yf.Ticker(ticker)
            hist = stock.history(
                start=try_date.strftime("%Y-%m-%d"),
                end=(try_date + timedelta(days=1)).strftime("%Y-%m-%d"),
            )
            if not hist.empty:
                close_price = hist['Close'].iloc[0]
                print(f"获取 {ticker} 于 {try_date} 的收盘价: {close_price}")
                return close_price
        except Exception as e:
            print(f"错误: 获取 {ticker} 于 {try_date} 收盘价失败，原因: {e}")
            continue
    print(f"未找到 {ticker} 在 {target_date} 附近的有效收盘价")
    return None

print(f"目标补齐日期: {target_date}, 基准日(3天后): {base_day}")

for sheet_name in wb.sheetnames:
    if not pattern.match(sheet_name):
        continue
    print(f"开始处理工作表: {sheet_name}")
    ws = wb[sheet_name]
    header = [cell.value for cell in ws[1]]
    required_cols = ["Date", "Ticker", "Previous Close", "3D Forward Change"]
    if any(col not in header for col in required_cols):
        print(f"工作表 {sheet_name} 缺少必要列，跳过")
        continue

    date_col = header.index("Date") + 1
    ticker_col = header.index("Ticker") + 1
    prev_close_col = header.index("Previous Close") + 1
    col_3d = header.index("3D Forward Change") + 1

    # 缓存基准日收盘价，避免重复请求
    base_day_close_cache = {}

    count_3d = 0

    for r in range(2, ws.max_row + 1):
        dt_cell = ws.cell(row=r, column=date_col).value
        if not dt_cell:
            continue
        if isinstance(dt_cell, datetime):
            dt_val = dt_cell.date()
        elif isinstance(dt_cell, str):
            dt_val = datetime.strptime(dt_cell[:10], "%Y-%m-%d").date()
        else:
            dt_val = dt_cell

        if dt_val != target_date:
            continue

        ticker = str(ws.cell(row=r, column=ticker_col).value).upper()
        prev_close = ws.cell(row=r, column=prev_close_col).value
        if prev_close is None or prev_close == 0:
            continue

        if ticker in base_day_close_cache:
            close_base = base_day_close_cache[ticker]
        else:
            close_base = get_price_realtime(ticker, base_day)
            if close_base is not None:
                base_day_close_cache[ticker] = close_base

        if close_base:
            change_3d = (close_base - prev_close) / prev_close
            ws.cell(row=r, column=col_3d).value = round(change_3d, 4)
            ws.cell(row=r, column=col_3d).number_format = "0.00%"
            count_3d += 1
            print(f"3D 补齐: 行 {r}，股票 {ticker}，日期 {dt_val}，涨跌幅 {change_3d:.4%}")

    print(f"工作表 {sheet_name} 补齐结果：3D共 {count_3d} 条")

wb.save(file_name)
print("指定日期3D Forward Change补齐完成")
