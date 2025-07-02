import pandas as pd
from datetime import datetime, timedelta
import os
import pytz
from openpyxl import load_workbook

# ==== 云端自动拉取最新 Excel ====
if "GITHUB_ACTIONS" in os.environ:
    os.system('rclone copy "gdrive:/Investing/Daily top options/option_activity_log.xlsx" ./ --drive-chunk-size 64M --progress --ignore-times')

# ✅ Toronto 时间
tz = pytz.timezone("America/Toronto")
now = datetime.now(tz)
today = now.date()
yesterday = today - timedelta(days=1)
target_day = today - timedelta(days=3)

file_name = "option_activity_log.xlsx"
wb = load_workbook(file_name, read_only=False, data_only=True)
sheet_names = wb.sheetnames

def is_month_sheet(name):
    try:
        datetime.strptime(name, "%Y-%m")
        return True
    except:
        return False

# ✅ Step 1：合并所有sheet用于匹配
df_all_list = []
for sheet_name in sheet_names:
    if is_month_sheet(sheet_name):
        df_tmp = pd.read_excel(file_name, sheet_name=sheet_name)
        df_tmp["SheetName"] = sheet_name
        df_all_list.append(df_tmp)
df_all = pd.concat(df_all_list, ignore_index=True)
df_all["Date"] = pd.to_datetime(df_all["Date"]).dt.date

fill_total = 0
for sheet_name in sheet_names:
    if not is_month_sheet(sheet_name):
        continue

    df = pd.read_excel(file_name, sheet_name=sheet_name)
    df["Date"] = pd.to_datetime(df["Date"]).dt.date

    if "3D Forward Change" not in df.columns:
        df["3D Forward Change"] = None

    # ✅ 只补今天-3天的数据，不回溯
    mask = df["Date"] == target_day
    fill_count = 0

    for i in df.index[mask]:
        row = df.loc[i]
        ticker = row["Ticker"]
        price_3ago = row["Previous Close"]

        match = df_all[(df_all["Ticker"] == ticker) & (df_all["Date"] == yesterday)]
        if not match.empty:
            price_3later = match.iloc[0]["Previous Close"]
            if price_3ago and price_3later and not pd.isna(price_3ago) and not pd.isna(price_3later):
                try:
                    change = (price_3later - price_3ago) / price_3ago
                    df.at[i, "3D Forward Change"] = round(change, 4)
                    fill_count += 1
                except Exception as e:
                    print(f"❌ Error for {ticker}: {e}")

    fill_total += fill_count

    with pd.ExcelWriter(file_name, mode="a", if_sheet_exists="replace", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"✅ Sheet {sheet_name}: 补齐日期 {target_day.strftime('%Y-%m-%d')} 共 {fill_count} 条 3D Forward Change")

print(f"✅ 所有 sheet 补齐完成，共 {fill_total} 条 3D Forward Change")

# ==== 云端上传 ====
if "GITHUB_ACTIONS" in os.environ:
    os.system('rclone copy ./option_activity_log.xlsx "gdrive:/Investing/Daily top options" --drive-chunk-size 64M --progress --ignore-times')
