import yfinance as yf
import pandas as pd
from datetime import datetime, timedelta
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import pytz
from math import log, sqrt, exp
from scipy.stats import norm
import requests



tz = pytz.timezone("America/Toronto")
now = datetime.now(tz)
today_str = now.strftime('%Y-%m-%d')
now_time_str = now.strftime('%H:%M')
month_sheet_name = now.strftime('%Y-%m')
file_name = "option_rank.xlsx"

# ==== 云端自动拉取最新 Excel ====
if "GITHUB_ACTIONS" in os.environ:
    os.system('rclone copy "gdrive:/Investing/Daily top options/option_rank.xlsx" ./ --drive-chunk-size 64M --progress --ignore-times')

# ==== 股票列表 ====
# === 标普500
url_sp500 = "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies"
headers = {"User-Agent": "Mozilla/5.0"}
html_sp500 = requests.get(url_sp500, headers=headers).text
sp500 = pd.read_html(html_sp500)[0]

# === 纳指100（自动识别正确表格）
url_nasdaq = "https://en.wikipedia.org/wiki/Nasdaq-100"
html_nasdaq = requests.get(url_nasdaq, headers=headers).text
nasdaq_tables = pd.read_html(html_nasdaq)

# 找到包含“TICKER”或“COMPANY”列的表
nasdaq = None
for df in nasdaq_tables:
    if any("ticker" in str(c).lower() or "symbol" in str(c).lower() for c in df.columns):
        nasdaq = df
        break
if nasdaq is None:
    raise ValueError(f"❌ 无法在 Nasdaq 页面中找到包含 ticker/symbol 的表格，实际表格数: {len(nasdaq_tables)}")


# === 合并
# === 自动识别列名 ===
def find_symbol_column(df):
    for c in df.columns:
        if 'symbol' in c.lower() or 'ticker' in c.lower():
            return c
    raise KeyError(f"找不到Symbol或Ticker列，实际列名: {list(df.columns)}")

sp500_symbol_col = find_symbol_column(sp500)
nasdaq_symbol_col = find_symbol_column(nasdaq)

tickers = list(set(sp500[sp500_symbol_col].astype(str).tolist() + nasdaq[nasdaq_symbol_col].astype(str).tolist()))
tickers = [t.replace('.', '-') for t in tickers]

# === 构建名称映射 ===
ticker_name_map = {}
for i in range(len(sp500)):
    try:
        symbol = str(sp500.loc[i, sp500_symbol_col]).replace('.', '-')
        name_col = next((c for c in sp500.columns if 'security' in c.lower() or 'company' in c.lower()), None)
        ticker_name_map[symbol] = sp500.loc[i, name_col] if name_col else ''
    except Exception:
        continue

for i in range(len(nasdaq)):
    try:
        symbol = str(nasdaq.loc[i, nasdaq_symbol_col]).replace('.', '-')
        if symbol not in ticker_name_map:
            name_col = next((c for c in nasdaq.columns if 'company' in c.lower()), None)
            ticker_name_map[symbol] = nasdaq.loc[i, name_col] if name_col else ''
    except Exception:
        continue

# ==== 抓取每只股票14天内所有期权，累计总成交量 ====
option_records = []
total_volume_dict = {}
option_detail_dict = {}
last_price_dict = {}

for ticker in tickers:
    print(f"Processing {ticker}")
    try:
        stock = yf.Ticker(ticker)
        # 抓当前价格
        try:
            close_price = stock.history(period="1d")["Close"].iloc[-1]
        except Exception:
            close_price = ""
        last_price_dict[ticker] = close_price


        try:
            expiries = stock.options[:4]  # ✅ 获取最近 4 个到期日
            expiry_this_week = expiries[:1]
            expiry_next_week = expiries[1:2]
            expiry_3rd_week = expiries[2:3]
            expiry_4th_week = expiries[3:4]
        except:
            continue  # 某些股票没有 options 字段或格式错误时跳过


        def fetch_option_block(expiry_list, label):
            all_opts = []
            for expiry in expiry_list:
                try:
                    chain = stock.option_chain(expiry)
                    calls = chain.calls.copy()
                    puts = chain.puts.copy()
                    calls["Type"] = "Call"
                    puts["Type"] = "Put"
                    calls["expiry"] = expiry
                    puts["expiry"] = expiry
                    calls["Group"] = label
                    puts["Group"] = label
                    all_opts.append(calls)
                    all_opts.append(puts)
                except:
                    continue
            return pd.concat(all_opts, ignore_index=True) if all_opts else None

        df1 = fetch_option_block(expiry_this_week, "This Week")
        df2 = fetch_option_block(expiry_next_week, "Next Week")
        df3 = fetch_option_block(expiry_3rd_week, "Third Week")
        df4 = fetch_option_block(expiry_4th_week, "Fourth Week")


        dfs = [df for df in [df1, df2, df3, df4] if df is not None]
        if not dfs:
            continue  # 无可用期权，跳过
        df_options = pd.concat(dfs, ignore_index=True)


        total_volume = df_options["volume"].sum()
        if total_volume > 0:
            total_volume_dict[ticker] = total_volume
            option_detail_dict[ticker] = df_options
    except Exception as e:
        continue


# ========== 希腊字母计算函数 ==========
def calc_greeks(option_type, S, K, T_days, IV, r=0.05):
    try:
        if S <= 0 or K <= 0 or T_days <= 0 or IV <= 0:
            return 0.0, 0.0, 0.0

        T = T_days / 365.0
        d1 = (log(S / K) + (r + 0.5 * IV**2) * T) / (IV * sqrt(T))
        d2 = d1 - IV * sqrt(T)

        if option_type == "Call":
            delta = norm.cdf(d1)
            theta = (-S * norm.pdf(d1) * IV / (2 * sqrt(T))
                     - r * K * exp(-r * T) * norm.cdf(d2))
        else:
            delta = -norm.cdf(-d1)
            theta = (-S * norm.pdf(d1) * IV / (2 * sqrt(T))
                     + r * K * exp(-r * T) * norm.cdf(-d2))

        gamma = norm.pdf(d1) / (S * IV * sqrt(T))
        return delta, gamma, theta / 365.0  # 每天的 Theta 损耗

    except:
        return 0.0, 0.0, 0.0

# ========== 分析成交量最高的前10只股票 ==========
top10 = sorted(total_volume_dict.items(), key=lambda x: -x[1])[:10]
records_raw = []

for ticker, _ in top10:
    company = ticker_name_map.get(ticker, '')
    close_price = last_price_dict.get(ticker, "")
    df_options = option_detail_dict[ticker]

    # ========== 分块处理 ==========
    for group_label in ["This Week", "Next Week", "Third Week", "Fourth Week"]:
        block = df_options[df_options["Group"] == group_label]
        if block.empty:
            continue

        top_block = block.sort_values("volume", ascending=False).head(10)

        for _, opt in top_block.iterrows():
            try:
                strike = float(opt["strike"])
                iv = float(opt["impliedVolatility"])
                expiry_date = pd.to_datetime(opt["expiry"])
                days_to_expiry = (expiry_date - datetime.now()).days

                delta, gamma, theta = calc_greeks(
                    opt["Type"],
                    float(close_price),
                    strike,
                    days_to_expiry,
                    iv
                )

                records_raw.append({
                    "Date": today_str,
                    "Time": now_time_str,
                    "Ticker": ticker,
                    "Last": round(float(close_price), 2),
                    "Type": opt["Type"],
                    "Strike": round(strike, 2),
                    "IV": round(iv * 100, 2),
                    "Volume": int(opt["volume"]),
                    "OI": int(opt["openInterest"]),
                    "Expiry": opt["expiry"],
                    "OptionLast": opt.get("lastPrice", "")
                })
            except Exception as e:
                print(f"⚠️ Error processing option for {ticker}: {e}")
        records_raw.append({})  # ✅ 分组间插空行



# ==== 写入 Excel（完全符合你的要求） ====
try:
    wb = load_workbook(file_name)
except FileNotFoundError:
    wb = Workbook()

if month_sheet_name in wb.sheetnames:
    ws = wb[month_sheet_name]
    first_write = False
else:
    ws = wb.create_sheet(title=month_sheet_name)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    first_write = True

headers = ["Date", "Time", "Ticker", "Last", "Type", "Strike", "IV", "Volume", "OI", "Expiry", "OptionLast"]


if first_write:
    ws.append(headers)
    ws.freeze_panes = "A2"
else:
    ws.append([])
    ws.append([])

for row in records_raw:
    if not row:
        ws.append([])
        continue
    ws.append([
        row.get("Date", ""), row.get("Time", ""), row.get("Ticker", ""), row.get("Last", ""),
        row.get("Type", ""), row.get("Strike", ""), row.get("IV", ""), row.get("Volume", ""), row.get("OI", ""),                
        row.get("Expiry", ""), row.get("OptionLast", "")
    ])

# 自动列宽
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2

wb.save(file_name)
# ✅ 提示用户运行结束
print("✅ option_rank 生成完毕，Excel 文件已保存：", file_name)


# ==== 云端同步回传 ====
if "GITHUB_ACTIONS" in os.environ:
    os.system('rclone copy ./option_rank.xlsx "gdrive:/Investing/Daily top options" --drive-chunk-size 64M --progress --ignore-times')