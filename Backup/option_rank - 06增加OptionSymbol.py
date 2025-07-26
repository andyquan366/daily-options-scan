import yfinance as yf
import pandas as pd
from datetime import datetime, timedelta
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import pytz

tz = pytz.timezone("America/Toronto")
now = datetime.now(tz)
today_str = now.strftime('%Y-%m-%d')
now_time_str = now.strftime('%H:%M')
month_sheet_name = now.strftime('%Y-%m')
file_name = "option_rank.xlsx"

# ==== 云端自动拉取最新 Excel ====
if "GITHUB_ACTIONS" in os.environ:
    os.system('rclone copy "gdrive:/Investing/Daily top options/option_rank.xlsx" ./ --drive-chunk-size 64M --progress --ignore-times')

# ==== 股票列表 ====
sp500 = pd.read_html('https://en.wikipedia.org/wiki/List_of_S%26P_500_companies')[0]
nasdaq = pd.read_html('https://en.wikipedia.org/wiki/Nasdaq-100')[4]
tickers = list(set(sp500['Symbol'].tolist() + nasdaq['Ticker'].tolist()))
tickers = [t.replace('.', '-') for t in tickers]
ticker_name_map = {}
for i in range(len(sp500)):
    symbol = sp500.loc[i, 'Symbol'].replace('.', '-')
    ticker_name_map[symbol] = sp500.loc[i, 'Security']
for i in range(len(nasdaq)):
    symbol = nasdaq.loc[i, 'Ticker'].replace('.', '-')
    if symbol not in ticker_name_map:
        ticker_name_map[symbol] = nasdaq.loc[i, 'Company']

# ==== 抓取每只股票14天内所有期权，累计总成交量 ====
option_records = []
total_volume_dict = {}
option_detail_dict = {}
last_price_dict = {}

for ticker in tickers:
    print(f"Processing {ticker}")
    try:
        stock = yf.Ticker(ticker)
        # 抓当前价格
        try:
            close_price = stock.history(period="1d")["Close"].iloc[-1]
        except Exception:
            close_price = ""
        last_price_dict[ticker] = close_price


        expiry_this_week = []
        expiry_next_week = []

        try:
            this_friday = now + timedelta(days=(4 - now.weekday()) % 7)  # 本周五
            for e in stock.options:
                exp_date = datetime.strptime(e, "%Y-%m-%d").date()
                days_diff = (exp_date - now.date()).days
                if 0 <= days_diff <= 14:
                    if exp_date <= this_friday.date():
                        expiry_this_week.append(e)
                    else:
                        expiry_next_week.append(e)
        except:
            continue  # 某些股票没有 options 字段或格式错误时跳过

        def fetch_option_block(expiry_list, label):
            all_opts = []
            for expiry in expiry_list:
                try:
                    chain = stock.option_chain(expiry)
                    calls = chain.calls.copy()
                    puts = chain.puts.copy()
                    calls["Type"] = "Call"
                    puts["Type"] = "Put"
                    calls["expiry"] = expiry
                    puts["expiry"] = expiry
                    calls["Group"] = label
                    puts["Group"] = label
                    all_opts.append(calls)
                    all_opts.append(puts)
                except:
                    continue
            return pd.concat(all_opts, ignore_index=True) if all_opts else None

        df1 = fetch_option_block(expiry_this_week, "This Week")
        df2 = fetch_option_block(expiry_next_week, "Next Week")

        if df1 is not None and df2 is not None:
            df_options = pd.concat([df1, df2], ignore_index=True)
        elif df1 is not None:
            df_options = df1
        elif df2 is not None:
            df_options = df2
        else:
            continue  # 无可用期权，跳过


        total_volume = df_options["volume"].sum()
        if total_volume > 0:
            total_volume_dict[ticker] = total_volume
            option_detail_dict[ticker] = df_options
    except Exception as e:
        continue

# ==== 选出成交量最大的前10只股票 ====
top10 = sorted(total_volume_dict.items(), key=lambda x: -x[1])[:10]
records_raw = []

for ticker, _ in top10:
    company = ticker_name_map.get(ticker, '')
    close_price = last_price_dict.get(ticker, "")
    df_options = option_detail_dict[ticker]
    # 按volume从大到小选10个

# ========== 分块处理 ==========
    for group_label in ["This Week", "Next Week"]:
        block = df_options[df_options["Group"] == group_label]
        if block.empty:
            continue
        top_block = block.sort_values("volume", ascending=False).head(10)

        for _, opt in top_block.iterrows():
            records_raw.append({
                "Date": today_str,
                "Time": now_time_str,
                "Ticker": ticker,
                "Company": company,
                "Last": round(float(close_price), 2),
                "Type": opt["Type"],
                "Strike": round(float(opt["strike"]), 2),
                "IV": round(opt["impliedVolatility"]*100, 2),
                "Volume": int(opt["volume"]),
                "OI": int(opt["openInterest"]),
                "Expiry": opt["expiry"],
                "OptionSymbol": opt["contractSymbol"]
            })
        records_raw.append({})  # 分组之间加空行


# ==== 写入 Excel（完全符合你的要求） ====
try:
    wb = load_workbook(file_name)
except FileNotFoundError:
    wb = Workbook()

if month_sheet_name in wb.sheetnames:
    ws = wb[month_sheet_name]
    first_write = False
else:
    ws = wb.create_sheet(title=month_sheet_name)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    first_write = True

headers = ["Date", "Time", "Ticker", "Company", "Last", "Type", "Strike", "IV", "Volume", "OI", "Expiry", "OptionSymbol"]

if first_write:
    ws.append(headers)
    ws.freeze_panes = "A2"
else:
    ws.append([])
    ws.append([])

for row in records_raw:
    if not row:
        ws.append([])
        continue
    ws.append([
        row.get("Date", ""), row.get("Time", ""), row.get("Ticker", ""), row.get("Company", ""), row.get("Last", ""),
        row.get("Type", ""), row.get("Strike", ""), row.get("IV", ""), row.get("Volume", ""), row.get("OI", ""), row.get("Expiry", ""), row.get("OptionSymbol", "")
    ])

# 自动列宽
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2

wb.save(file_name)
# ✅ 提示用户运行结束
print("✅ option_rank 生成完毕，Excel 文件已保存：", file_name)


# ==== 云端同步回传 ====
if "GITHUB_ACTIONS" in os.environ:
    os.system('rclone copy ./option_rank.xlsx "gdrive:/Investing/Daily top options" --drive-chunk-size 64M --progress --ignore-times')
