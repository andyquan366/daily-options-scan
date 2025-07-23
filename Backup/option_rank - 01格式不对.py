import yfinance as yf
import pandas as pd
from datetime import datetime, timedelta
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import pytz

tz = pytz.timezone("America/Toronto")
now = datetime.now(tz)
today_str = now.strftime('%Y-%m-%d')
now_time_str = now.strftime('%H:%M')
month_sheet_name = now.strftime('%Y-%m')
file_name = "option_rank.xlsx"

# ==== 云端自动拉取最新 Excel ====
if "GITHUB_ACTIONS" in os.environ:
    os.system('rclone copy "gdrive:/Investing/Daily top options/option_rank.xlsx" ./ --drive-chunk-size 64M --progress --ignore-times')

# ==== 股票列表（请换成你自己的全名单源） ====
sp500 = pd.read_html('https://en.wikipedia.org/wiki/List_of_S%26P_500_companies')[0]
nasdaq = pd.read_html('https://en.wikipedia.org/wiki/Nasdaq-100')[4]
tickers = list(set(sp500['Symbol'].tolist() + nasdaq['Ticker'].tolist()))
tickers = [t.replace('.', '-') for t in tickers]
ticker_name_map = {}
for i in range(len(sp500)):
    symbol = sp500.loc[i, 'Symbol'].replace('.', '-')
    ticker_name_map[symbol] = sp500.loc[i, 'Security']
for i in range(len(nasdaq)):
    symbol = nasdaq.loc[i, 'Ticker'].replace('.', '-')
    if symbol not in ticker_name_map:
        ticker_name_map[symbol] = nasdaq.loc[i, 'Company']

# ==== 抓取每只股票14天内所有期权，累计总成交量 ====
option_records = []
total_volume_dict = {}
option_detail_dict = {}

for ticker in tickers:
    print(f"Processing {ticker}")  # 每处理一只股票先打印进度
    try:
        stock = yf.Ticker(ticker)
        expiry_dates = [e for e in stock.options if (datetime.strptime(e, "%Y-%m-%d").date() - now.date()).days <= 14]
        all_options = []
        for expiry in expiry_dates:
            try:
                chain = stock.option_chain(expiry)
                calls = chain.calls.copy()
                puts = chain.puts.copy()
                calls["Type"] = "Call"
                puts["Type"] = "Put"
                calls["expiry"] = expiry
                puts["expiry"] = expiry
                all_options.append(calls)
                all_options.append(puts)
            except:
                continue
        if not all_options:
            continue
        df_options = pd.concat(all_options, ignore_index=True)
        df_options = df_options[["Type", "strike", "impliedVolatility", "volume", "openInterest", "expiry", "contractSymbol"]]
        total_volume = df_options["volume"].sum()
        if total_volume > 0:
            total_volume_dict[ticker] = total_volume
            option_detail_dict[ticker] = df_options
    except Exception as e:
        continue

# ==== 选出成交量最大的前10只股票 ====
top10 = sorted(total_volume_dict.items(), key=lambda x: -x[1])[:10]
records_raw = []

for ticker, _ in top10:
    company = ticker_name_map.get(ticker, '')
    df_options = option_detail_dict[ticker]
    # 按volume从大到小选10个
    top_options = df_options.sort_values("volume", ascending=False).head(10)
    for _, opt in top_options.iterrows():
        records_raw.append({
            "Date": today_str,
            "Time": now_time_str,
            "Ticker": ticker,
            "Company": company,
            "Type": opt["Type"],
            "Strike": opt["strike"],
            "IV": round(opt["impliedVolatility"]*100, 2) if pd.notna(opt["impliedVolatility"]) else '',
            "Volume": int(opt["volume"]),
            "OI": int(opt["openInterest"]),
            "Expiry": opt["expiry"]
        })
    # 加空行分块
    records_raw.append({})

# ==== 写入 Excel（每月sheet，每次运行两空行，表头，分股票空行，自动列宽）====
try:
    wb = load_workbook(file_name)
except FileNotFoundError:
    wb = Workbook()
if month_sheet_name in wb.sheetnames:
    ws = wb[month_sheet_name]
else:
    ws = wb.create_sheet(title=month_sheet_name)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

# 每次运行加两空行
ws.append([])
ws.append([])
headers = ["Date", "Time", "Ticker", "Company", "Type", "Strike", "IV", "Volume", "OI", "Expiry"]
ws.append(headers)

for row in records_raw:
    if not row:  # 空行分块
        ws.append([])
        continue
    ws.append([
        row.get("Date", ""), row.get("Time", ""), row.get("Ticker", ""), row.get("Company", ""),
        row.get("Type", ""), row.get("Strike", ""), row.get("IV", ""), row.get("Volume", ""), row.get("OI", ""), row.get("Expiry", "")
    ])

# 自动列宽
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2

wb.save(file_name)

# ==== 云端同步回传 ====
if "GITHUB_ACTIONS" in os.environ:
    os.system('rclone copy ./option_rank.xlsx "gdrive:/Investing/Daily top options" --drive-chunk-size 64M --progress --ignore-times')
