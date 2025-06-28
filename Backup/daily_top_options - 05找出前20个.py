import yfinance as yf
import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

print("📅 获取股票列表...")
sp500 = pd.read_html('https://en.wikipedia.org/wiki/List_of_S%26P_500_companies')[0]
nasdaq = pd.read_html('https://en.wikipedia.org/wiki/Nasdaq-100')[4]
tickers = list(set(sp500['Symbol'].tolist() + nasdaq['Ticker'].tolist()))
tickers = [t.replace('.', '-') for t in tickers]

records = []
today_str = datetime.now().strftime('%Y-%m-%d')
print(f"🔍 分析期权活跃度（共 {len(tickers)} 支股票）...")

for ticker in tickers:
    try:
        stock = yf.Ticker(ticker)
        expiry_dates = stock.options
        if not expiry_dates:
            continue

        all_calls = []
        all_puts = []

        for expiry in expiry_dates:
            try:
                chain = stock.option_chain(expiry)
                calls = chain.calls.copy()
                puts = chain.puts.copy()
                calls['expiry'] = expiry
                puts['expiry'] = expiry
                all_calls.append(calls)
                all_puts.append(puts)
            except:
                continue

        if not all_calls or not all_puts:
            continue

        merged_calls = pd.concat(all_calls, ignore_index=True)
        merged_puts = pd.concat(all_puts, ignore_index=True)

        call_total_volume = merged_calls['volume'].sum()
        put_total_volume = merged_puts['volume'].sum()
        if call_total_volume == 0:
            continue
        put_call_ratio = round(put_total_volume / call_total_volume, 2)

        top_call = merged_calls.sort_values("volume", ascending=False).iloc[0]
        if top_call['openInterest'] > 0 and top_call['volume'] > 0:
            call_vo_ratio = round(top_call['volume'] / top_call['openInterest'], 2)
        else:
            continue

        top_put = merged_puts.sort_values("volume", ascending=False).iloc[0]
        if top_put['openInterest'] > 0 and top_put['volume'] > 0:
            put_vo_ratio = round(top_put['volume'] / top_put['openInterest'], 2)
        else:
            continue

        iv_skew = round(top_call['impliedVolatility'] * 100 - top_put['impliedVolatility'] * 100, 2)

        if put_call_ratio < 0.8 and call_vo_ratio > put_vo_ratio and iv_skew > 0:
            sentiment = "Bullish"
        elif put_call_ratio > 1.2 and put_vo_ratio > call_vo_ratio and iv_skew < 0:
            sentiment = "Bearish"
        else:
            sentiment = "Neutral"

        records.append({
            'Date': today_str,
            'Ticker': ticker,
            'Type': 'Call',
            'Strike': top_call['strike'],
            'IV': round(top_call['impliedVolatility'] * 100, 2),
            'Volume': int(top_call['volume']),
            'Open Interest': int(top_call['openInterest']),
            'Volume/OI': call_vo_ratio,
            'Expiry': top_call['expiry'],
            'Put/Call Ratio': put_call_ratio,
            'Call V/OI': call_vo_ratio,
            'Put V/OI': put_vo_ratio,
            'IV Skew': iv_skew,
            'Sentiment': sentiment
        })

        records.append({
            'Date': today_str,
            'Ticker': ticker,
            'Type': 'Put',
            'Strike': top_put['strike'],
            'IV': round(top_put['impliedVolatility'] * 100, 2),
            'Volume': int(top_put['volume']),
            'Open Interest': int(top_put['openInterest']),
            'Volume/OI': put_vo_ratio,
            'Expiry': top_put['expiry'],
            'Put/Call Ratio': put_call_ratio,
            'Call V/OI': call_vo_ratio,
            'Put V/OI': put_vo_ratio,
            'IV Skew': iv_skew,
            'Sentiment': sentiment
        })

    except Exception:
        continue

if not records:
    print("❌ 未找到合适的期权活跃记录")
    exit()

df = pd.DataFrame(records)

# ✅ 提取 Call 记录，按 Volume/OI 降序，选出前 20 个 ticker
call_df = df[df['Type'] == 'Call'].copy()
top20_tickers = (
    call_df.sort_values('Volume/OI', ascending=False)
           .drop_duplicates(subset='Ticker')
           .head(20)['Ticker']
           .tolist()
)

# ✅ 回头从总表中取出这 20 个 ticker 的 Call 和 Put 记录
df = df[df['Ticker'].isin(top20_tickers)]

# ✅ 排序逻辑：Ticker + Type（Call先） + Volume/OI 降序
df['TypeRank'] = df['Type'].map({'Call': 0, 'Put': 1})
df = df.sort_values(by=['Ticker', 'TypeRank', 'Volume/OI'], ascending=[True, True, False])
df = df.drop(columns=['TypeRank'])

print(f"\n📊 {today_str} 最强20个股票的期权活跃数据:")
print(df.to_string(index=False))

# ✅ 统计情绪分布
sentiment_counts = df['Sentiment'].value_counts()
bullish = sentiment_counts.get("Bullish", 0)
bearish = sentiment_counts.get("Bearish", 0)
neutral = sentiment_counts.get("Neutral", 0)
print(f"\n🧠 今日情绪分布：Bullish {bullish} / Bearish {bearish} / Neutral {neutral}")

file_name = "option_activity_log.xlsx"
if not os.path.exists(file_name):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Top Options"
    for r in dataframe_to_rows(df, index=False, header=True):
        ws1.append(r)
    ws1.append([])
    ws1.append([])

    ws2 = wb.create_sheet(title="Sentiment Stats")
    ws2.append(["Date", "Bullish", "Bearish", "Neutral"])
    ws2.append([today_str, bullish, bearish, neutral])
else:
    wb = load_workbook(file_name)
    ws1 = wb[wb.sheetnames[0]]
    ws1.append([])
    ws1.append([])
    for r in dataframe_to_rows(df, index=False, header=False):
        ws1.append(r)

    if "Sentiment Stats" in wb.sheetnames:
        ws2 = wb["Sentiment Stats"]
    else:
        ws2 = wb.create_sheet(title="Sentiment Stats")
        ws2.append(["Date", "Bullish", "Bearish", "Neutral"])
    ws2.append([today_str, bullish, bearish, neutral])

# ✅ 情绪字段上色
sentiment_col_index = None
for i, cell in enumerate(ws1[1], start=1):
    if cell.value == "Sentiment":
        sentiment_col_index = i
        break

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

if sentiment_col_index:
    for row in ws1.iter_rows(min_row=2, min_col=sentiment_col_index, max_col=sentiment_col_index):
        for cell in row:
            if cell.value == "Bullish":
                cell.fill = green_fill
            elif cell.value == "Bearish":
                cell.fill = red_fill
            elif cell.value == "Neutral":
                cell.fill = yellow_fill

wb.save(file_name)
print(f"✅ 已追加今日数据并保存至: {file_name}")
