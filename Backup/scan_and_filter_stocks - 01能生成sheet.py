import pandas as pd
from datetime import datetime, timedelta
import yfinance as yf
from openpyxl import load_workbook
import re
import pytz
import os

# ==== 云端自动拉取最新 Excel ====
if "GITHUB_ACTIONS" in os.environ:
    os.system('rclone copy "gdrive:/Investing/Daily top options/option_activity_log.xlsx" ./ --drive-chunk-size 64M --progress --ignore-times')

file_name = "option_activity_log.xlsx"
wb = load_workbook(file_name)

# 设置基准日为昨天
today = datetime.today().date()
base_day = today - timedelta(days=1)

pattern = re.compile(r"^\d{4}-\d{2}$")  # 匹配 yyyy-mm 格式

# 新增：定义筛选条件的函数
def filter_stocks(ws, date_col, ticker_col, price_change_col, change_7d_col):
    filtered_stocks = []
    for r in range(2, ws.max_row + 1):
        dt_cell = ws.cell(row=r, column=date_col).value
        if not dt_cell:
            continue
        if isinstance(dt_cell, datetime):
            dt_val = dt_cell.date()
        elif isinstance(dt_cell, str):
            dt_val = datetime.strptime(dt_cell[:10], "%Y-%m-%d").date()
        else:
            dt_val = dt_cell

        ticker = str(ws.cell(row=r, column=ticker_col).value).upper()
        price_change = ws.cell(row=r, column=price_change_col).value
        change_7d = ws.cell(row=r, column=change_7d_col).value

        # 筛选条件：Price Change 为负数，7D Change 为正数
        if price_change is not None and change_7d is not None:
            if price_change < 0 and change_7d > 0:
                filtered_stocks.append({
                    'Date': dt_val,
                    'Ticker': ticker,
                    'Price Change': price_change,
                    '7D Change': change_7d
                })
    return filtered_stocks

# 处理所有工作表
for sheet_name in wb.sheetnames:
    if not pattern.match(sheet_name):
        continue
    print(f"开始处理工作表: {sheet_name}")
    ws = wb[sheet_name]
    header = [cell.value for cell in ws[1]]

    # 确保列存在
    required_cols = ["Date", "Ticker", "Price Change", "7D Change"]
    if any(col not in header for col in required_cols):
        print(f"工作表 {sheet_name} 缺少必要列，跳过")
        continue

    # 获取列索引
    date_col = header.index("Date") + 1
    ticker_col = header.index("Ticker") + 1
    price_change_col = header.index("Price Change") + 1
    change_7d_col = header.index("7D Change") + 1

    # 筛选符合条件的数据
    filtered_stocks = filter_stocks(ws, date_col, ticker_col, price_change_col, change_7d_col)

    if filtered_stocks:
        # 新建工作表保存筛选结果
        new_sheet_name = f"Filtered_{sheet_name}"
        new_ws = wb.create_sheet(new_sheet_name)

        # 写入表头
        new_ws.append(['Date', 'Ticker', 'Price Change', '7D Change'])

        # 写入筛选后的数据
        for stock in filtered_stocks:
            new_ws.append([stock['Date'], stock['Ticker'], stock['Price Change'], stock['7D Change']])

        print(f"工作表 {sheet_name} 中符合条件的数据已写入新的工作表 {new_sheet_name}")

# 保存更新后的 Excel 文件
wb.save(file_name)
print("数据筛选并保存完成")

if "GITHUB_ACTIONS" in os.environ:
    os.system('rclone copy ./option_activity_log.xlsx "gdrive:/Investing/Daily top options" --drive-chunk-size 64M --progress --ignore-times')
