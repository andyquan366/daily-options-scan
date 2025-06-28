import yfinance as yf
import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

# 获取股票列表
print("📅 获取股票列表...")
sp500 = pd.read_html('https://en.wikipedia.org/wiki/List_of_S%26P_500_companies')[0]
nasdaq = pd.read_html('https://en.wikipedia.org/wiki/Nasdaq-100')[4]
tickers = list(set(sp500['Symbol'].tolist() + nasdaq['Ticker'].tolist()))
tickers = [t.replace('.', '-') for t in tickers]

records = []
print("🔍 分析期权活跃度...")

today_str = datetime.now().strftime('%Y-%m-%d')

for ticker in tickers[:200]:
    try:
        stock = yf.Ticker(ticker)
        expiry_dates = stock.options
        if not expiry_dates:
            continue

        expiry = expiry_dates[0]
        chain = stock.option_chain(expiry)
        calls = chain.calls
        puts = chain.puts

        if calls.empty or puts.empty:
            continue

        # 总成交量（计算 Put/Call Ratio）
        call_total_volume = calls['volume'].sum()
        put_total_volume = puts['volume'].sum()
        if call_total_volume == 0:
            continue
        put_call_ratio = round(put_total_volume / call_total_volume, 2)

        # 找出最活跃的 call 和 put
        top_call = calls.sort_values("volume", ascending=False).iloc[0]
        top_put = puts.sort_values("volume", ascending=False).iloc[0]

        if top_call['openInterest'] == 0 or top_put['openInterest'] == 0:
            continue

        call_vo_ratio = round(top_call['volume'] / top_call['openInterest'], 2)
        put_vo_ratio = round(top_put['volume'] / top_put['openInterest'], 2)

        iv_skew = round(top_call['impliedVolatility'] * 100 - top_put['impliedVolatility'] * 100, 2)

        # call 记录
        records.append({
            'Date': today_str,
            'Ticker': ticker,
            'Type': 'Call',
            'Strike': top_call['strike'],
            'IV': round(top_call['impliedVolatility'] * 100, 2),
            'Volume': int(top_call['volume']),
            'Open Interest': int(top_call['openInterest']),
            'Volume/OI': call_vo_ratio,
            'Expiry': expiry,
            'Put/Call Ratio': put_call_ratio,
            'Call V/OI': call_vo_ratio,
            'Put V/OI': put_vo_ratio,
            'IV Skew': iv_skew
        })

        # put 记录
        records.append({
            'Date': today_str,
            'Ticker': ticker,
            'Type': 'Put',
            'Strike': top_put['strike'],
            'IV': round(top_put['impliedVolatility'] * 100, 2),
            'Volume': int(top_put['volume']),
            'Open Interest': int(top_put['openInterest']),
            'Volume/OI': put_vo_ratio,
            'Expiry': expiry,
            'Put/Call Ratio': put_call_ratio,
            'Call V/OI': call_vo_ratio,
            'Put V/OI': put_vo_ratio,
            'IV Skew': iv_skew
        })

    except Exception:
        continue

# 根据 Volume/OI 排序，只保留前 20
if not records:
    print("❌ 未找到合适的期权活跃记录")
    exit()

df = pd.DataFrame(records)
df = df.sort_values('Volume/OI', ascending=False).head(20)

print(f"\n📊 {today_str} 最活跃 Call/Put 期权前20:")
print(df.to_string(index=False))

# 写入 Excel
file_name = "option_activity_log.xlsx"
if not os.path.exists(file_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    ws.append([])
    ws.append([])
    wb.save(file_name)
    print(f"✅ 创建文件并保存: {file_name}")
else:
    wb = load_workbook(file_name)
    ws = wb.active
    ws.append([])
    ws.append([])
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)
    wb.save(file_name)
    print(f"✅ 已附加今日数据到: {file_name}")
