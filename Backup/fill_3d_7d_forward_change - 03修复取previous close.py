import pandas as pd
from datetime import datetime, timedelta
import yfinance as yf
from openpyxl import load_workbook
import re
import pytz
import os

# ==== 云端自动拉取最新 Excel ====
if "GITHUB_ACTIONS" in os.environ:
    os.system('rclone copy "gdrive:/Investing/Daily top options/option_activity_log.xlsx" ./ --drive-chunk-size 64M --progress --ignore-times')

file_name = "option_activity_log.xlsx"
wb = load_workbook(file_name)

# 设置基准日为昨天
today = datetime.today().date()
base_day = today - timedelta(days=1)

pattern = re.compile(r"^\d{4}-\d{2}$")  # 匹配 yyyy-mm 格式

def get_previous_close_yf(ticker, target_date):
    try:
        stock = yf.Ticker(ticker)
        info = stock.info
        return info.get("previousClose", None)
    except Exception as e:
        print(f"错误: 获取 {ticker} 的 Previous Close 失败，原因: {e}")
        return None

print(f"基准日（昨天）为: {base_day}")

for sheet_name in wb.sheetnames:
    if not pattern.match(sheet_name):
        continue
    print(f"开始处理工作表: {sheet_name}")
    ws = wb[sheet_name]

    header = [cell.value for cell in ws[1]]

# 如果缺列则自动添加列头
    if "3D Forward Change" not in header:
        ws.cell(row=1, column=len(header) + 1).value = "3D Forward Change"
        header.append("3D Forward Change")
    if "7D Forward Change" not in header:
        ws.cell(row=1, column=len(header) + 1).value = "7D Forward Change"
        header.append("7D Forward Change")

# 更新列索引（注意 +1，因为 openpyxl 列是从1开始）
    date_col = header.index("Date") + 1
    ticker_col = header.index("Ticker") + 1
    prev_close_col = header.index("Previous Close") + 1
    col_3d = header.index("3D Forward Change") + 1
    col_7d = header.index("7D Forward Change") + 1


    # 先获取昨天所有ticker的收盘价，缓存起来，避免多次请求
    yesterday_close_cache = {}

    # 遍历行，缓存昨天日期的所有ticker及其previous close
    for r in range(2, ws.max_row + 1):
        dt_cell = ws.cell(row=r, column=date_col).value
        if not dt_cell:
            continue
        if isinstance(dt_cell, datetime):
            dt_val = dt_cell.date()
        elif isinstance(dt_cell, str):
            dt_val = datetime.strptime(dt_cell[:10], "%Y-%m-%d").date()
        else:
            dt_val = dt_cell
        if dt_val == base_day:
            ticker = str(ws.cell(row=r, column=ticker_col).value).upper()
            prev_close = ws.cell(row=r, column=prev_close_col).value
            if prev_close is not None:
                yesterday_close_cache[ticker] = prev_close

    print(f"基准日收盘价缓存数: {len(yesterday_close_cache)}")

    count_3d = 0
    count_7d = 0

    # 再遍历行，找日期为 base_day -3 天和 base_day -7 天，补齐对应涨跌幅
    for r in range(2, ws.max_row + 1):
        dt_cell = ws.cell(row=r, column=date_col).value
        if not dt_cell:
            continue
        if isinstance(dt_cell, datetime):
            dt_val = dt_cell.date()
        elif isinstance(dt_cell, str):
            dt_val = datetime.strptime(dt_cell[:10], "%Y-%m-%d").date()
        else:
            dt_val = dt_cell

        ticker = str(ws.cell(row=r, column=ticker_col).value).upper()
        prev_close = ws.cell(row=r, column=prev_close_col).value
        if prev_close is None or prev_close == 0:
            continue

        days_diff_3 = (base_day - dt_val).days
        days_diff_7 = (base_day - dt_val).days

        # 3D补齐逻辑
        if days_diff_3 == 3:
            # 从缓存中找昨天的close价
            if ticker in yesterday_close_cache:
                close_yesterday = yesterday_close_cache[ticker]
            else:
                close_yesterday = get_previous_close_yf(ticker, base_day)
                if close_yesterday is not None:
                    yesterday_close_cache[ticker] = close_yesterday
            if close_yesterday:
                change_3d = (close_yesterday - prev_close) / prev_close
                ws.cell(row=r, column=col_3d).value = round(change_3d, 4)
                ws.cell(row=r, column=col_3d).number_format = "0.00%"
                count_3d += 1
                print(f"3D 补齐: 行 {r}，股票 {ticker}，日期 {dt_val}，涨跌幅 {change_3d:.4%}")

        # 7D补齐逻辑
        if days_diff_7 == 7:
            # 从缓存中找昨天的close价
            if ticker in yesterday_close_cache:
                close_yesterday = yesterday_close_cache[ticker]
            else:
                close_yesterday = get_previous_close_yf(ticker, base_day)
                if close_yesterday is not None:
                    yesterday_close_cache[ticker] = close_yesterday
            if close_yesterday:
                change_7d = (close_yesterday - prev_close) / prev_close
                ws.cell(row=r, column=col_7d).value = round(change_7d, 4)
                ws.cell(row=r, column=col_7d).number_format = "0.00%"
                count_7d += 1
                print(f"7D 补齐: 行 {r}，股票 {ticker}，日期 {dt_val}，涨跌幅 {change_7d:.4%}")

    print(f"工作表 {sheet_name} 补齐结果：3D共 {count_3d} 条，7D共 {count_7d} 条")

wb.save(file_name)
print("补齐完成")

if "GITHUB_ACTIONS" in os.environ:
    os.system('rclone copy ./option_activity_log.xlsx "gdrive:/Investing/Daily top options" --drive-chunk-size 64M --progress --ignore-times')