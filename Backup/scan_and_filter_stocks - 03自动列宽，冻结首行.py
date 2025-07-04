from datetime import datetime
from openpyxl import load_workbook
import re
import os
from openpyxl.utils import get_column_letter

if "GITHUB_ACTIONS" in os.environ:
    os.system('rclone copy "gdrive:/Investing/Daily top options/option_activity_log.xlsx" ./ --drive-chunk-size 64M --progress --ignore-times')

file_name = "option_activity_log.xlsx"
wb = load_workbook(file_name)
pattern = re.compile(r"^\d{4}-\d{2}$")

def filter_stocks(ws, date_col, ticker_col, company_col,
                  price_change_col, change_7d_col, change_3d_forward_col, change_7d_forward_col):
    filtered_stocks = []
    seen = set()
    for r in range(2, ws.max_row + 1):
        dt_cell = ws.cell(row=r, column=date_col).value
        if not dt_cell:
            continue
        if isinstance(dt_cell, datetime):
            dt_val = dt_cell.date()
        elif isinstance(dt_cell, str):
            dt_val = datetime.strptime(dt_cell[:10], "%Y-%m-%d").date()
        else:
            dt_val = dt_cell

        ticker = str(ws.cell(row=r, column=ticker_col).value).upper()
        company = ws.cell(row=r, column=company_col).value
        price_change = ws.cell(row=r, column=price_change_col).value
        change_7d = ws.cell(row=r, column=change_7d_col).value
        change_3d_forward = ws.cell(row=r, column=change_3d_forward_col).value
        change_7d_forward = ws.cell(row=r, column=change_7d_forward_col).value

        key = (dt_val, ticker, company)
        if key in seen:
            continue
        seen.add(key)

        # 只用 Price Change <0 和 7D Change >0 作为筛选条件
        if price_change is not None and change_7d is not None:
            if price_change < 0 and change_7d > 0:
                filtered_stocks.append({
                    'Date': dt_val,
                    'Ticker': ticker,
                    'Company': company,
                    'Price Change': price_change,
                    '7D Change': change_7d,
                    '3D Forward Change': change_3d_forward,
                    '7D Forward Change': change_7d_forward,
                })
    return filtered_stocks



def auto_adjust_column_width(ws):
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column_cells:
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        adjusted_width = max_length + 1
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = adjusted_width


for sheet_name in wb.sheetnames:
    if not pattern.match(sheet_name):
        continue
    print(f"开始处理工作表: {sheet_name}")
    ws = wb[sheet_name]
    header = [cell.value for cell in ws[1]]

    # 根据您提供的列名完整校验
    required_cols = ["Date", "Ticker", "Company",
                     "Price Change", "7D Change", "3D Forward Change", "7D Forward Change"]
    if any(col not in header for col in required_cols):
        print(f"工作表 {sheet_name} 缺少必要列，跳过")
        continue

    date_col = header.index("Date") + 1
    ticker_col = header.index("Ticker") + 1
    company_col = header.index("Company") + 1
    price_change_col = header.index("Price Change") + 1
    change_7d_col = header.index("7D Change") + 1
    change_3d_forward_col = header.index("3D Forward Change") + 1
    change_7d_forward_col = header.index("7D Forward Change") + 1

    filtered_stocks = filter_stocks(ws, date_col, ticker_col, company_col,
                                   price_change_col, change_7d_col, change_3d_forward_col, change_7d_forward_col)
    print(f"工作表 {sheet_name} 筛选到符合条件的数据数量: {len(filtered_stocks)}")

    if filtered_stocks:
        new_sheet_name = f"Filtered_{sheet_name}"
        if new_sheet_name in wb.sheetnames:
            wb.remove(wb[new_sheet_name])
        new_ws = wb.create_sheet(new_sheet_name)

# 冻结首行
        new_ws.freeze_panes = 'A2'


        # 写入表头，顺序严格按您说的来
        new_ws.append(['Date', 'Ticker', 'Company', 'Price Change', '7D Change', '3D Forward Change', '7D Forward Change'])
        row_idx = 2  # 初始化行号，从第二行开始写入数据
        for stock in filtered_stocks:
            new_ws.append([
                stock['Date'],
                stock['Ticker'],
                stock['Company'],
                stock['Price Change'],
                stock['7D Change'],
                stock['3D Forward Change'],
                stock['7D Forward Change'],
            ])
            for col in range(4, 8):
                new_ws.cell(row=row_idx, column=col).number_format = '0.00%'
            row_idx += 1

# 调用自动列宽函数
        auto_adjust_column_width(new_ws)

        print(f"写入 {len(filtered_stocks)} 条数据到新工作表 {new_sheet_name}")

wb.save(file_name)
print("数据筛选并保存完成")
