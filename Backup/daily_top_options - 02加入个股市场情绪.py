import yfinance as yf
import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

# 获取股票列表
print("📅 获取股票列表...")
sp500 = pd.read_html('https://en.wikipedia.org/wiki/List_of_S%26P_500_companies')[0]
nasdaq = pd.read_html('https://en.wikipedia.org/wiki/Nasdaq-100')[4]
tickers = list(set(sp500['Symbol'].tolist() + nasdaq['Ticker'].tolist()))
tickers = [t.replace('.', '-') for t in tickers]

records = []
print("🔍 分析期权活跃度 (全部到期日)...")

today_str = datetime.now().strftime('%Y-%m-%d')

for ticker in tickers[:200]:
    try:
        stock = yf.Ticker(ticker)
        expiry_dates = stock.options
        if not expiry_dates:
            continue

        all_calls = []
        all_puts = []

        for expiry in expiry_dates:
            try:
                chain = stock.option_chain(expiry)
                calls = chain.calls.copy()
                puts = chain.puts.copy()
                calls['expiry'] = expiry
                puts['expiry'] = expiry
                all_calls.append(calls)
                all_puts.append(puts)
            except:
                continue

        if not all_calls or not all_puts:
            continue

        merged_calls = pd.concat(all_calls, ignore_index=True)
        merged_puts = pd.concat(all_puts, ignore_index=True)

        call_total_volume = merged_calls['volume'].sum()
        put_total_volume = merged_puts['volume'].sum()
        if call_total_volume == 0:
            continue
        put_call_ratio = round(put_total_volume / call_total_volume, 2)

        top_call = merged_calls.sort_values("volume", ascending=False).iloc[0]
        if top_call['openInterest'] > 0 and top_call['volume'] > 0:
            call_vo_ratio = round(top_call['volume'] / top_call['openInterest'], 2)
        else:
            continue

        top_put = merged_puts.sort_values("volume", ascending=False).iloc[0]
        if top_put['openInterest'] > 0 and top_put['volume'] > 0:
            put_vo_ratio = round(top_put['volume'] / top_put['openInterest'], 2)
        else:
            continue

        iv_skew = round(top_call['impliedVolatility'] * 100 - top_put['impliedVolatility'] * 100, 2)

        # 判断情绪
        if put_call_ratio < 0.8 and call_vo_ratio > put_vo_ratio and iv_skew > 0:
            sentiment = "Bullish"
        elif put_call_ratio > 1.2 and put_vo_ratio > call_vo_ratio and iv_skew < 0:
            sentiment = "Bearish"
        else:
            sentiment = "Neutral"

        records.append({
            'Date': today_str,
            'Ticker': ticker,
            'Type': 'Call',
            'Strike': top_call['strike'],
            'IV': round(top_call['impliedVolatility'] * 100, 2),
            'Volume': int(top_call['volume']),
            'Open Interest': int(top_call['openInterest']),
            'Volume/OI': call_vo_ratio,
            'Expiry': top_call['expiry'],
            'Put/Call Ratio': put_call_ratio,
            'Call V/OI': call_vo_ratio,
            'Put V/OI': put_vo_ratio,
            'IV Skew': iv_skew,
            'Sentiment': sentiment
        })

        records.append({
            'Date': today_str,
            'Ticker': ticker,
            'Type': 'Put',
            'Strike': top_put['strike'],
            'IV': round(top_put['impliedVolatility'] * 100, 2),
            'Volume': int(top_put['volume']),
            'Open Interest': int(top_put['openInterest']),
            'Volume/OI': put_vo_ratio,
            'Expiry': top_put['expiry'],
            'Put/Call Ratio': put_call_ratio,
            'Call V/OI': call_vo_ratio,
            'Put V/OI': put_vo_ratio,
            'IV Skew': iv_skew,
            'Sentiment': sentiment
        })

    except Exception:
        continue

if not records:
    print("❌ 未找到合适的期权活跃记录")
    exit()

df = pd.DataFrame(records)
df = df.sort_values('Volume/OI', ascending=False).head(20)

print(f"\n📊 {today_str} 总期活跃排行 TOP20:")
print(df.to_string(index=False))

file_name = "option_activity_log.xlsx"
if not os.path.exists(file_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    ws.append([])
    ws.append([])
    wb.save(file_name)
    print(f"✅ 创建文件并保存: {file_name}")
else:
    wb = load_workbook(file_name)
    ws = wb.active
    ws.append([])
    ws.append([])
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)
    wb.save(file_name)
    print(f"✅ 已附加今日数据到: {file_name}")
