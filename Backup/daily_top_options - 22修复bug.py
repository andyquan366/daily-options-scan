import yfinance as yf
import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image as XLImage
import pytz
from datetime import datetime


print('📅 获取股票列表...')
sp500 = pd.read_html('https://en.wikipedia.org/wiki/List_of_S%26P_500_companies')[0]
nasdaq = pd.read_html('https://en.wikipedia.org/wiki/Nasdaq-100')[4]
tickers = list(set(sp500['Symbol'].tolist() + nasdaq['Ticker'].tolist()))
tickers = [t.replace('.', '-') for t in tickers]

ticker_name_map = {}
for i in range(len(sp500)):
    symbol = sp500.loc[i, 'Symbol'].replace('.', '-')
    ticker_name_map[symbol] = sp500.loc[i, 'Security']
for i in range(len(nasdaq)):
    symbol = nasdaq.loc[i, 'Ticker'].replace('.', '-')
    if symbol not in ticker_name_map:
        ticker_name_map[symbol] = nasdaq.loc[i, 'Company']


# ✅ 全局设定 Toronto 本地时间，所有时间统一用 now，不可重复定义
tz = pytz.timezone("America/Toronto")
now = datetime.now(tz)
today = now.date()
today_str = now.strftime('%Y-%m-%d')   # e.g., '2025-06-28'
now_time_str = now.strftime('%H:%M')   # e.g., '21:03'

records = []
volume_summary = {}


for ticker in tickers:
    print(f'▶ 正在处理 {ticker} ...')
    try:
        stock = yf.Ticker(ticker)
        expiry_dates = stock.options


# ✅ 计算涨跌幅（百分比）
        hist = stock.history(period='2d')
        if len(hist) >= 2:
            prev_close = hist['Close'].iloc[-2]
            last_price = hist['Close'].iloc[-1]
            price_change = round((last_price - prev_close) / prev_close * 100, 2)
            price_change_str = f"{price_change:+.2f}%"
        else:
            price_change_str = "N/A"

        expiry_dates = [e for e in expiry_dates if (datetime.strptime(e, "%Y-%m-%d").date() - today).days <= 10]
        if not expiry_dates:
            continue

        all_calls, all_puts = [], []
        for expiry in expiry_dates:
            try:
                chain = stock.option_chain(expiry)
                calls, puts = chain.calls.copy(), chain.puts.copy()
                calls['expiry'], puts['expiry'] = expiry, expiry
                all_calls.append(calls)
                all_puts.append(puts)
            except:
                continue

        if not all_calls or not all_puts:
            continue

        merged_calls = pd.concat(all_calls, ignore_index=True)
        merged_puts = pd.concat(all_puts, ignore_index=True)
        if merged_calls.empty or merged_puts.empty:
            continue

        top_call = merged_calls.sort_values('volume', ascending=False).iloc[0]
        top_put = merged_puts.sort_values('volume', ascending=False).iloc[0]

        call_volume = top_call['volume']
        put_volume = top_put['volume']
        total_volume = merged_calls['volume'].sum() + merged_puts['volume'].sum()

        if total_volume < 3000:
            continue

        put_call_ratio = round(put_volume / call_volume, 4) if call_volume != 0 else float('inf')
        premium_skew = round(top_call['lastPrice'] - top_put['lastPrice'], 2)
        volume_diff_ratio = (call_volume - put_volume) / (call_volume + put_volume) if (call_volume + put_volume) != 0 else 0

# ✅ Premium Skew 打分（满分 50）
        if premium_skew >= 7:
            score_premium = 50
        elif premium_skew >= 4:
            score_premium = 40
        elif premium_skew >= 1:
            score_premium = 30
        elif premium_skew >= -2:
            score_premium = 20
        elif premium_skew >= -5:
            score_premium = 10
        else:
            score_premium = 0

# ✅ Volume Diff Ratio 打分（满分 30）
        if volume_diff_ratio >= 0.8:
            score_vol = 30
        elif volume_diff_ratio >= 0.4:
            score_vol = 24
        elif volume_diff_ratio >= 0:
            score_vol = 18
        elif volume_diff_ratio >= -0.4:
            score_vol = 12
        elif volume_diff_ratio >= -0.8:
            score_vol = 6
        else:
            score_vol = 0

# ✅ Put/Call Ratio 打分（满分 20）
        if put_call_ratio <= 0.2:
            score_pcr = 20
        elif put_call_ratio <= 0.8:
            score_pcr = 15
        elif put_call_ratio <= 2:
            score_pcr = 10
        elif put_call_ratio <= 4:
            score_pcr = 5
        else:
            score_pcr = 0



        total_score = score_pcr + score_vol + score_premium

        sentiment = (
            "Strong Bullish" if total_score >= 80 else
            "Bullish" if total_score >= 60 else
            "Neutral" if total_score >= 40 else
            "Bearish" if total_score >= 20 else
            "Strong Bearish"
        )

        for option_type, top_option in zip(['Call', 'Put'], [top_call, top_put]):
            records.append({
                'Date': today_str,'Time': now_time_str,'Ticker': ticker, 'Company': ticker_name_map.get(ticker, ''),
                'Type': option_type, 'Strike': top_option['strike'],
                'IV': round(top_option['impliedVolatility'] * 100, 2),
                'Volume': int(top_option['volume']), 'Expiry': top_option['expiry'],
                'Put/Call Ratio': put_call_ratio, 'Premium Skew': premium_skew,
                'Volume Diff Ratio': round(volume_diff_ratio, 4), 'Score': total_score,
                'Sentiment': sentiment, 'Contract Symbol': top_option['contractSymbol'],
                'Price Change': price_change_str})

        volume_summary[ticker] = total_volume

    except Exception as e:
        continue

df = pd.DataFrame(records)
if df.empty:
    exit()

top40_tickers = sorted(volume_summary.items(), key=lambda x: x[1], reverse=True)[:40]
top40_set = set(t[0] for t in top40_tickers)
df = df[df['Ticker'].isin(top40_set)].copy()
df['TypeRank'] = df['Type'].apply(lambda x: 0 if x == 'Call' else 1)
df = df.sort_values(by=['Score', 'Ticker', 'TypeRank'], ascending=[False, True, True])
df.drop(columns=['TypeRank'], inplace=True)  # ✅ 排序后立即删掉，避免导出


sentiments = ["Strong Bullish", "Bullish", "Neutral", "Bearish", "Strong Bearish"]
sentiment_counts = df['Sentiment'].value_counts()

# ✅ 写入 Excel
file_name = "option_activity_log.xlsx"
month_sheet_name = now.strftime("%Y-%m")
year_sheet_name = now.strftime("%Y")

if not os.path.exists(file_name):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = month_sheet_name   # 改这里，改为年月名
    ws1.freeze_panes = 'A2'
    for r in dataframe_to_rows(df, index=False, header=True):
        ws1.append(r)
    ws1.append([])
    ws1.append([])
    ws2 = wb.create_sheet(title=year_sheet_name)  # 改这里，改为年份名
    ws2.append(["Date", "Time","Strong Bullish", "Bullish", "Neutral", "Bearish", "Strong Bearish", "Score"])
else:
    wb = load_workbook(file_name)
    if month_sheet_name in wb.sheetnames:
        ws1 = wb[month_sheet_name]
    else:
        ws1 = wb.create_sheet(month_sheet_name)
    ws1.freeze_panes = 'A2'
    ws1.append([])
    ws1.append([])
    for r in dataframe_to_rows(df, index=False, header=False):
        ws1.append(r)
    if year_sheet_name in wb.sheetnames:
        ws2 = wb[year_sheet_name]
    else:
        ws2 = wb.create_sheet(year_sheet_name)
    ws2.freeze_panes = 'A2'



# 检查表头是否存在，不存在时添加
if ws2.max_row == 0:
    ws2.append(["Date", "Time", "Strong Bullish", "Bullish", "Neutral", "Bearish", "Strong Bearish", "Score"])

sentiments = ["Strong Bullish", "Bullish", "Neutral", "Bearish", "Strong Bearish"]
ticker_sentiment = df[['Ticker', 'Sentiment']].drop_duplicates(subset=['Ticker'])
sentiment_counts = ticker_sentiment['Sentiment'].value_counts()

sentiment_score = (
    sentiment_counts.get("Strong Bullish", 0) * 2 +
    sentiment_counts.get("Bullish", 0) * 1 +
    sentiment_counts.get("Neutral", 0) * 0 +
    sentiment_counts.get("Bearish", 0) * (-1) +
    sentiment_counts.get("Strong Bearish", 0) * (-2)
)


last_data_row = ws2.max_row
if last_data_row >= 2:
    last_date = ws2.cell(row=last_data_row, column=1).value  # 第1列是 Date
    if isinstance(last_date, str) and last_date != now.strftime("%Y-%m-%d"):
        ws2.append([])
        ws2.append([])
ws2.append([
    now.strftime("%Y-%m-%d"),
    now.strftime("%H:%M"),
    sentiment_counts.get("Strong Bullish", 0),
    sentiment_counts.get("Bullish", 0),
    sentiment_counts.get("Neutral", 0),
    sentiment_counts.get("Bearish", 0),
    sentiment_counts.get("Strong Bearish", 0),
    sentiment_score,
])

# ✅ 着色
fills = {
    "Strong Bullish": "C6EFCE",
    "Bullish": "C6EFCE",
    "Neutral": "FFEB9C",
    "Bearish": "FFC7CE",
    "Strong Bearish": "FFC7CE"
}

sentiment_col_index = [c.value for c in ws1[1]].index("Sentiment") + 1
for row in ws1.iter_rows(min_row=2, min_col=sentiment_col_index, max_col=sentiment_col_index):
    for cell in row:
        cell.fill = PatternFill(start_color=fills.get(cell.value, "FFFFFF"), fill_type="solid")

# ✅ 自动调整列宽（对两个工作表都执行）
from openpyxl.utils import get_column_letter

for ws in [ws1, ws2]:
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 1

wb.save(file_name)

# 重新加载工作簿，确保数据已写入
wb = load_workbook(file_name)
ws2 = wb[year_sheet_name]

# ✅ 给第二页 Score 列添加红绿背景，表示分数变化（只对比最近两行）
score_col_index = [c.value for c in ws2[1]].index("Score") + 1
for row in range(3, ws2.max_row + 1):  # 从第3行开始（第1行为表头，第2行为第一条数据）
    prev_score = ws2.cell(row=row - 1, column=score_col_index).value
    curr_score = ws2.cell(row=row, column=score_col_index).value
    if isinstance(prev_score, int) and isinstance(curr_score, int):
        if curr_score > prev_score:
            ws2.cell(row=row, column=score_col_index).fill = PatternFill(start_color="C6EFCE", fill_type="solid")  # 绿
        elif curr_score < prev_score:
            ws2.cell(row=row, column=score_col_index).fill = PatternFill(start_color="FFC7CE", fill_type="solid")  # 红


wb.save(file_name)
print(f"✅ 数据已保存至 {file_name}")


# 读取文件画图
df_all = pd.read_excel(file_name, sheet_name=year_sheet_name)
df_all = df_all[df_all['Date'].notna()]
df_all['Date'] = pd.to_datetime(df_all['Date'])
df_all = df_all.sort_values('Date')

wb = load_workbook(file_name)
ws2 = wb[year_sheet_name]

ws2._images.clear()

max_col = ws2.max_column
start_col = max_col + 2

row_offset = 0
img_height_rows = 20

unique_dates = sorted(df_all['Date'].dt.date.unique())

for date_val in unique_dates:
    day_data = df_all[df_all['Date'].dt.date == date_val]
    day_data = day_data.sort_values('Time')
    x = day_data['Time']
    y = day_data['Score']

    plt.figure(figsize=(10,3))
    plt.plot(x, y, marker='o', linestyle='-')
    plt.title(f'Score on {date_val}')
    plt.xlabel('Time')
    plt.ylabel('Score')
    plt.xticks(x, x, rotation=45)
    plt.tight_layout()

    img_file = f'score_{date_val}.png'
    plt.savefig(img_file)
    plt.close()

    img = XLImage(img_file)
    img.anchor = f"{get_column_letter(start_col)}{1 + row_offset}"
    ws2.add_image(img)
    row_offset += img_height_rows



wb.save(file_name)  # 画图完毕后保存
print("✅ 数据和图形都已保存完成")