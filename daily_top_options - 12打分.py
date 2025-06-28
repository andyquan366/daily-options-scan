import yfinance as yf
import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

print('📅 获取股票列表...')
sp500 = pd.read_html('https://en.wikipedia.org/wiki/List_of_S%26P_500_companies')[0]
nasdaq = pd.read_html('https://en.wikipedia.org/wiki/Nasdaq-100')[4]
tickers = list(set(sp500['Symbol'].tolist() + nasdaq['Ticker'].tolist()))
tickers = [t.replace('.', '-') for t in tickers]

ticker_name_map = {}
for i in range(len(sp500)):
    symbol = sp500.loc[i, 'Symbol'].replace('.', '-')
    ticker_name_map[symbol] = sp500.loc[i, 'Security']
for i in range(len(nasdaq)):
    symbol = nasdaq.loc[i, 'Ticker'].replace('.', '-')
    if symbol not in ticker_name_map:
        ticker_name_map[symbol] = nasdaq.loc[i, 'Company']

records = []
volume_summary = {}
today_str = datetime.now().strftime('%Y-%m-%d')

for ticker in tickers:
    print(f'▶ 正在处理 {ticker} ...')
    try:
        stock = yf.Ticker(ticker)
        expiry_dates = stock.options
        today = datetime.today()
        expiry_dates = [e for e in expiry_dates if (datetime.strptime(e, "%Y-%m-%d") - today).days <= 10]
        if not expiry_dates:
            continue

        all_calls, all_puts = [], []
        for expiry in expiry_dates:
            try:
                chain = stock.option_chain(expiry)
                calls, puts = chain.calls.copy(), chain.puts.copy()
                calls['expiry'], puts['expiry'] = expiry, expiry
                all_calls.append(calls)
                all_puts.append(puts)
            except:
                continue

        if not all_calls or not all_puts:
            continue

        merged_calls = pd.concat(all_calls, ignore_index=True)
        merged_puts = pd.concat(all_puts, ignore_index=True)
        if merged_calls.empty or merged_puts.empty:
            continue

        top_call = merged_calls.sort_values('volume', ascending=False).iloc[0]
        top_put = merged_puts.sort_values('volume', ascending=False).iloc[0]

        call_volume = top_call['volume']
        put_volume = top_put['volume']
        total_volume = merged_calls['volume'].sum() + merged_puts['volume'].sum()

        if total_volume < 3000:
            continue

        put_call_ratio = round(put_volume / call_volume, 4) if call_volume != 0 else float('inf')
        iv_skew = round(top_call['impliedVolatility'] * 100 - top_put['impliedVolatility'] * 100, 4)
        volume_diff_ratio = (call_volume - put_volume) / (call_volume + put_volume) if (call_volume + put_volume) != 0 else 0





        score_pcr = 40 if put_call_ratio < 0.4 else 30 if put_call_ratio < 0.6 else 20 if put_call_ratio < 0.8 else 10 if put_call_ratio < 1.0 else 0
        score_skew = 30 if iv_skew > 5 else 20 if iv_skew > 2 else 10 if iv_skew > -2 else 5 if iv_skew > -5 else 0
        score_vol = 30 if volume_diff_ratio > 0.7 else 20 if volume_diff_ratio > 0.4 else 10 if volume_diff_ratio > 0.1 else 5 if volume_diff_ratio > -0.1 else 0

        total_score = score_pcr + score_skew + score_vol

        sentiment = (
            "Strong Bullish" if total_score >= 85 else
            "Bullish" if total_score >= 65 else
            "Neutral" if total_score >= 35 else
            "Bearish" if total_score >= 15 else
            "Strong Bearish"
        )




        for option_type, top_option in zip(['Call', 'Put'], [top_call, top_put]):
            records.append({
                'Date': today_str, 'Ticker': ticker, 'Company': ticker_name_map.get(ticker, ''),
                'Type': option_type, 'Strike': top_option['strike'],
                'IV': round(top_option['impliedVolatility'] * 100, 2),
                'Volume': int(top_option['volume']), 'Expiry': top_option['expiry'],
                'Put/Call Ratio': put_call_ratio, 'IV Skew': iv_skew,
                'Volume Diff Ratio': round(volume_diff_ratio, 4), 'Score': total_score,
                'Sentiment': sentiment, 'contractSymbol': top_option['contractSymbol']})

        volume_summary[ticker] = total_volume

    except Exception as e:
        print(f'⚠️ 跳过 {ticker}，错误：{e}')
        continue

df = pd.DataFrame(records)
if df.empty:
    print('❌ 最终 records 为空，没有生成有效数据')
    exit()

top40_tickers = sorted(volume_summary.items(), key=lambda x: x[1], reverse=True)[:40]
top40_set = set(t[0] for t in top40_tickers)
df = df[df['Ticker'].isin(top40_set)]
df = df.sort_values(by=['Score'], ascending=False)

sentiment_counts = df['Sentiment'].value_counts()
bullish = sentiment_counts.get("Bullish", 0) + sentiment_counts.get("Strong Bullish", 0)
bearish = sentiment_counts.get("Bearish", 0) + sentiment_counts.get("Strong Bearish", 0)
neutral = sentiment_counts.get("Neutral", 0)
print(f"🧠 今日情绪分布：Bullish {bullish} / Bearish {bearish} / Neutral {neutral}")

# 后续Excel写入代码与第一版本完全一致（略）


# 写入 Excel
file_name = "option_activity_log.xlsx"
if not os.path.exists(file_name):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Top Options"
    ws1.freeze_panes = 'A2'  # 冻结第一行
else:
    wb = load_workbook(file_name)
    ws1 = wb["Top Options"]

# 清空旧数据但保留表头
if ws1.max_row > 1:
    ws1.delete_rows(2, ws1.max_row - 1)

for r in dataframe_to_rows(df, index=False, header=True):
    ws1.append(r)

# 添加情绪分布统计页
if "Sentiment Stats" in wb.sheetnames:
    ws2 = wb["Sentiment Stats"]
else:
    ws2 = wb.create_sheet(title="Sentiment Stats")

ws2.append(["Date", "Bullish", "Bearish", "Neutral"])
ws2.append([today_str, bullish, bearish, neutral])

# 着色规则
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

# 找 Sentiment 列索引
sentiment_col_index = None
for idx, cell in enumerate(ws1[1], start=1):
    if cell.value == "Sentiment":
        sentiment_col_index = idx
        break

if sentiment_col_index:
    for row in ws1.iter_rows(min_row=2, min_col=sentiment_col_index, max_col=sentiment_col_index):
        for cell in row:
            if cell.value in ("Bullish", "Strong Bullish"):
                cell.fill = green_fill
            elif cell.value in ("Bearish", "Strong Bearish"):
                cell.fill = red_fill
            elif cell.value == "Neutral":
                cell.fill = yellow_fill

# 自动调整列宽
for col in ws1.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws1.column_dimensions[col_letter].width = max_length + 1

wb.save(file_name)
print(f"✅ 数据已保存至 {file_name}")
