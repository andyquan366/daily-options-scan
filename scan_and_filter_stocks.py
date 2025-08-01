from openpyxl import load_workbook
import re
import os
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from collections import defaultdict, OrderedDict
from openpyxl.styles import PatternFill
import pandas as pd

scan_start_date = datetime(2025, 6, 27).date()

if "GITHUB_ACTIONS" in os.environ:
    os.system('rclone copy "gdrive:/Investing/Daily top options/option_activity_log.xlsx" ./ --drive-chunk-size 64M --progress --ignore-times')

file_name = "option_activity_log.xlsx"
wb = load_workbook(file_name)
pattern = re.compile(r"^\d{4}-\d{2}$")

def get_daily_change_yf(ticker, ref_date):
    import yfinance as yf
    from datetime import timedelta
    stock = yf.Ticker(ticker)
    max_lookback = 10
    for i in range(max_lookback):
        day = ref_date - timedelta(days=i)
        hist = stock.history(start=day.strftime('%Y-%m-%d'), end=(day + timedelta(days=1)).strftime('%Y-%m-%d'))
        if not hist.empty:
            close = hist['Close'].iloc[0]
            prev_day = day - timedelta(days=1)
            prev_hist = stock.history(start=prev_day.strftime('%Y-%m-%d'), end=day.strftime('%Y-%m-%d'))
            if not prev_hist.empty:
                prev_close = prev_hist['Close'].iloc[0]
                return round((close - prev_close) / prev_close, 6)
    return None

def get_7d_change_yf(ticker, ref_date):
    import yfinance as yf
    from datetime import timedelta
    stock = yf.Ticker(ticker)
    max_lookback = 10
    for i in range(max_lookback):
        end_day = ref_date - timedelta(days=i)
        start_day = end_day - timedelta(days=7)
        hist = stock.history(start=start_day.strftime('%Y-%m-%d'), end=(end_day + timedelta(days=1)).strftime('%Y-%m-%d'))
        if len(hist) >= 2:
            start_close = hist['Close'].iloc[0]
            end_close = hist['Close'].iloc[-1]
            return round((end_close - start_close) / start_close, 6)
    return None

def filter_stocks(sheet_name, scan_start_date):
    df = pd.read_excel("option_activity_log.xlsx", sheet_name=sheet_name)
    df['Date'] = pd.to_datetime(df['Date']).dt.date
    filtered_df = df[df['Date'] >= scan_start_date]

    records_dict = {}
    for _, row in filtered_df.iterrows():
        dt_val = row['Date']
        ticker = str(row['Ticker']).upper()
        company = row['Company']
        price_change = row['Price Change']
        change_7d = row['7D Change']
        change_3d_forward = row['3D Forward Change']
        change_7d_forward = row['7D Forward Change']
        score = row['Score']
        try:
            score = float(score)
        except:
            score = 0.0

        key = (dt_val, ticker, company)
        records_dict[key] = {
            'Date': dt_val,
            'Ticker': ticker,
            'Company': company,
            'Price Change': price_change,
            '7D Change': change_7d,
            '3D Forward Change': change_3d_forward,
            '7D Forward Change': change_7d_forward,
            'Score': score,
        }

    filtered_stocks = list(records_dict.values())
    score_accum = {}
    for stock in filtered_stocks:
        score_key = (stock['Date'], stock['Ticker'])
        if score_key not in score_accum:
            score_accum[score_key] = []
        score_accum[score_key].append(stock['Score'])

    for stock in filtered_stocks:
        score_key = (stock['Date'], stock['Ticker'])
        scores = score_accum.get(score_key, [])
        avg_score = sum(scores) / len(scores) if scores else 0.0
        stock['AVG Score'] = avg_score

    return filtered_stocks




def safe_average(lst):
    filtered = [x for x in lst if x is not None]
    if filtered:
        return sum(filtered) / len(filtered)
    return None

def write_stats(ws, date, d3_dict, d7_dict, score_dict, start_row):
    combos = [('+','+'), ('+','-'), ('-','+'), ('-','-')]

    col_sign1 = 10  # J列
    col_sign2 = 11  # K列
    col_avg_3d = 12 # L列
    col_avg_7d = 13 # M列
    col_avg_score = 14 # N列

    for i, combo in enumerate(combos):
        avg_3d = safe_average(d3_dict[date][combo])
        avg_7d = safe_average(d7_dict[date][combo])
        avg_score = safe_average(score_dict[date][combo])

        row = start_row + i  # 从主数据第一行开始写统计连续4行

        ws.cell(row=row, column=col_sign1, value=combo[0])
        ws.cell(row=row, column=col_sign2, value=combo[1])
        ws.cell(row=row, column=col_avg_3d, value=avg_3d)
        ws.cell(row=row, column=col_avg_7d, value=avg_7d)
        ws.cell(row=row, column=col_avg_score, value=int(round(avg_score)) if avg_score is not None else None)

        if avg_3d is not None:
            ws.cell(row=row, column=col_avg_3d).number_format = '0.00%'
        if avg_7d is not None:
            ws.cell(row=row, column=col_avg_7d).number_format = '0.00%'

def auto_adjust_column_width(ws):
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column_cells:
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        adjusted_width = max_length + 1
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = adjusted_width

daily_3d = defaultdict(lambda: defaultdict(list))
daily_7d = defaultdict(lambda: defaultdict(list))
daily_score = defaultdict(lambda: defaultdict(list))

current_year = datetime.now().year
yearly_sheet_name = f"Filtered_{current_year}"
# 每次都重新生成 Filtered_2025 表（不保留旧内容）
if yearly_sheet_name in wb.sheetnames:
    del wb[yearly_sheet_name]
yearly_ws = wb.create_sheet(yearly_sheet_name)
yearly_ws.append(['Date', 'Ticker', 'Company', 'AVG Score', 'Price Change', '7D Change', '3D Forward Change', '7D Forward Change'])
yearly_ws.freeze_panes = 'C2'
last_yearly_date = None  # 用于追踪上一个写入

for sheet_name in wb.sheetnames:
    if not pattern.match(sheet_name):
        continue
    print(f"开始处理工作表: {sheet_name}")


    filtered_stocks = filter_stocks(sheet_name, scan_start_date)

    print(f"工作表 {sheet_name} 筛选到符合条件的数据数量: {len(filtered_stocks)}")

    ws = wb[sheet_name]


    # 先写统计表头（第一行）
    new_sheet_name = f"Filtered_{sheet_name}"
    if new_sheet_name in wb.sheetnames:
        wb.remove(wb[new_sheet_name])
    new_ws = wb.create_sheet(new_sheet_name)
    new_ws.freeze_panes = 'C2'

    # 主表头
    new_ws.append(['Date', 'Ticker', 'Company', 'AVG Score', 'Price Change', '7D Change', '3D Forward Change', '7D Forward Change'])
    # 统计表头写第一行右侧固定列
    new_ws.cell(row=1, column=10, value="Price Change")
    new_ws.cell(row=1, column=11, value="7D Change")
    new_ws.cell(row=1, column=12, value="3D Forward Change")
    new_ws.cell(row=1, column=13, value="7D Forward Change")
    new_ws.cell(row=1, column=14, value="AVG Score")

    prev_date = None
    row_idx = 2
    date_rows = []

    filtered_stocks.sort(key=lambda x: (x['Date'], -x['AVG Score']))

    for stock in filtered_stocks:
        curr_date = stock['Date']

    # 计算符号组合
        sign_price = '+' if stock['Price Change'] >= 0 else '-'
        sign_7d = '+' if stock['7D Change'] >= 0 else '-'
        combo = (sign_price, sign_7d)

    # 往统计字典里添加数据
        daily_3d[curr_date][combo].append(stock['3D Forward Change'])
        daily_7d[curr_date][combo].append(stock['7D Forward Change'])
        daily_score[curr_date][combo].append(stock['Score'])

        if curr_date != prev_date and prev_date is not None:
            # 写统计，传入该日期第一条数据行号，统计写入连续4行
            write_stats(new_ws, prev_date, daily_3d, daily_7d, daily_score, min(date_rows))
            # 日期块之间空2行
            row_idx += 2
            date_rows.clear()

        new_ws.cell(row=row_idx, column=1, value=curr_date)
        new_ws.cell(row=row_idx, column=2, value=stock['Ticker'])
        new_ws.cell(row=row_idx, column=3, value=stock['Company'])
        new_ws.cell(row=row_idx, column=4, value=stock['AVG Score'])
        new_ws.cell(row=row_idx, column=5, value=stock['Price Change'])
        new_ws.cell(row=row_idx, column=6, value=stock['7D Change'])
        new_ws.cell(row=row_idx, column=7, value=stock['3D Forward Change'])
        new_ws.cell(row=row_idx, column=8, value=stock['7D Forward Change'])

        for col in range(5, 9):
            new_ws.cell(row=row_idx, column=col).number_format = '0.00%'


        highlight_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        pc = stock['Price Change']
        d7 = stock['7D Change']

# ✅ 原始标红条件：
# 如果 Price Change 和 7D Change 都有值，且 7D Change > 0，
# 且 Price Change / 7D Change 的比例在 [-0.7, -0.2] 之间，就标红
#        if pc is not None and d7 is not None and d7 > 0:
#            ratio = pc / d7
#            if -0.7 <= ratio <= -0.2:


# ✅ 现在标红条件：分数 ≥80 或 ≤20，且 Price Change ≥ 0
        score = stock['AVG Score']
        if (score >= 80 or score <= 20) and pc is not None and pc >= 0:


                for col in range(1, 9):  # 把第1列到第8列全部标红
                    new_ws.cell(row=row_idx, column=col).fill = highlight_fill

                if last_yearly_date is not None and curr_date != last_yearly_date:
                    yearly_ws.append([])
                    yearly_ws.append([])

        # ✅ 同时追加到 Filtered_2025
                yearly_ws.append([
                    curr_date,
                    stock['Ticker'],
                    stock['Company'],
                    stock['AVG Score'],
                    pc,
                    d7,
                    stock['3D Forward Change'],
                    stock['7D Forward Change']
                ])
                for c in range(5, 9):
                    yearly_ws.cell(row=yearly_ws.max_row, column=c).number_format = '0.00%'

# 更新日期
                last_yearly_date = curr_date

        date_rows.append(row_idx)
        prev_date = curr_date
        row_idx += 1

    # 写最后一天的统计
    if prev_date is not None and date_rows:
        write_stats(new_ws, prev_date, daily_3d, daily_7d, daily_score, min(date_rows))

    auto_adjust_column_width(new_ws)

    new_ws.column_dimensions['I'].width = 8

auto_adjust_column_width(yearly_ws)
yearly_ws.append([''] * 8)  # 添加空白行，防止被标签遮挡
if yearly_sheet_name in wb.sheetnames:
    wb._sheets.insert(0, wb._sheets.pop(wb.sheetnames.index(yearly_sheet_name)))


wb.save(file_name)

if "GITHUB_ACTIONS" in os.environ:
    os.system('rclone copy ./option_activity_log.xlsx "gdrive:/Investing/Daily top options" --drive-chunk-size 64M --progress --ignore-times')
