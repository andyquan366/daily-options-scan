import pandas as pd
from datetime import datetime, timedelta
import yfinance as yf
from openpyxl import load_workbook
import re
import pytz
import os

# ==== 云端自动拉取最新 Excel ====
if "GITHUB_ACTIONS" in os.environ:
    os.system('rclone copy "gdrive:/Investing/Daily top options/option_activity_log.xlsx" ./ --drive-chunk-size 64M --progress --ignore-times')

file_name = "option_activity_log.xlsx"
wb = load_workbook(file_name)

# 统一用 Toronto 时区时间，避免本地云端时间差异
tz = pytz.timezone("America/Toronto")
now = datetime.now(tz)
today = now.date()
base_day = today - timedelta(days=1)

pattern = re.compile(r"^\d{4}-\d{2}$")  # 匹配 yyyy-mm 格式

def get_price_realtime(ticker, target_date, max_lookback=14):
    """
    往前最多回溯 max_lookback 天，跳过周末，找最近有效交易日收盘价
    """
    from datetime import timedelta
    for i in range(max_lookback):
        try_date = target_date - timedelta(days=i)
        if try_date.weekday() >= 5:  # 跳过周末
            continue
        try:
            stock = yf.Ticker(ticker)
            hist = stock.history(
                start=try_date.strftime("%Y-%m-%d"),
                end=(try_date + timedelta(days=1)).strftime("%Y-%m-%d"),
                # 移除 progress 参数，避免报错
            )
            if not hist.empty:
                close_price = hist['Close'].iloc[0]
                print(f"获取 {ticker} 于 {try_date} 的收盘价: {close_price}")
                return close_price
        except Exception as e:
            print(f"错误: 获取 {ticker} 于 {try_date} 收盘价失败，原因: {e}")
            continue
    print(f"未找到 {ticker} 在 {target_date} 附近的有效收盘价")
    return None

print(f"基准日（昨天）为: {base_day}")

for sheet_name in wb.sheetnames:
    if not pattern.match(sheet_name):
        continue
    print(f"开始处理工作表: {sheet_name}")
    ws = wb[sheet_name]
    header = [str(cell.value).strip() if cell.value is not None else '' for cell in ws[1]]
    # 加打印列标题
    print(f"工作表 {sheet_name} 的第一行列标题是: {header}")

    required_cols = ["Date", "Ticker", "Previous Close", "3D Forward Change", "7D Forward Change"]
    
    missing_cols = [col for col in required_cols if col not in header]
    if missing_cols:
        print(f"缺失的列: {missing_cols}")
        print(f"工作表 {sheet_name} 缺少必要列，跳过")
        continue

    date_col = header.index("Date") + 1
    ticker_col = header.index("Ticker") + 1
    prev_close_col = header.index("Previous Close") + 1
    col_3d = header.index("3D Forward Change") + 1
    col_7d = header.index("7D Forward Change") + 1

    # 先获取昨天所有ticker的收盘价，缓存起来，避免多次请求
    yesterday_close_cache = {}

    # 遍历行，缓存昨天日期的所有ticker及其previous close
    for r in range(2, ws.max_row + 1):
        dt_cell = ws.cell(row=r, column=date_col).value
        if not dt_cell:
            continue
        if isinstance(dt_cell, datetime):
            dt_val = dt_cell.date()
        elif isinstance(dt_cell, str):
            dt_val = datetime.strptime(dt_cell[:10], "%Y-%m-%d").date()
        else:
            dt_val = dt_cell
        if dt_val == base_day:
            ticker = str(ws.cell(row=r, column=ticker_col).value).upper()
            prev_close = ws.cell(row=r, column=prev_close_col).value
            if prev_close is not None:
                yesterday_close_cache[ticker] = prev_close

    print(f"基准日收盘价缓存数: {len(yesterday_close_cache)}")

    count_3d = 0
    count_7d = 0

    # 再遍历行，找日期为 base_day -3 天和 base_day -7 天，补齐对应涨跌幅
    for r in range(2, ws.max_row + 1):
        dt_cell = ws.cell(row=r, column=date_col).value
        if not dt_cell:
            continue
        if isinstance(dt_cell, datetime):
            dt_val = dt_cell.date()
        elif isinstance(dt_cell, str):
            dt_val = datetime.strptime(dt_cell[:10], "%Y-%m-%d").date()
        else:
            dt_val = dt_cell

        ticker = str(ws.cell(row=r, column=ticker_col).value).upper()
        prev_close = ws.cell(row=r, column=prev_close_col).value
        if prev_close is None or prev_close == 0:
            continue

        days_diff_3 = (base_day - dt_val).days
        days_diff_7 = (base_day - dt_val).days

        # 3D补齐逻辑
        if days_diff_3 == 3:
            # 从缓存中找昨天的close价
            if ticker in yesterday_close_cache:
                close_yesterday = yesterday_close_cache[ticker]
            else:
                close_yesterday = get_price_realtime(ticker, base_day)
                if close_yesterday is not None:
                    yesterday_close_cache[ticker] = close_yesterday
            if close_yesterday:
                change_3d = (close_yesterday - prev_close) / prev_close
                ws.cell(row=r, column=col_3d).value = round(change_3d, 4)
                ws.cell(row=r, column=col_3d).number_format = "0.00%"
                count_3d += 1
                print(f"3D 补齐: 行 {r}，股票 {ticker}，日期 {dt_val}，涨跌幅 {change_3d:.4%}")

        # 7D补齐逻辑
        if days_diff_7 == 7:
            # 从缓存中找昨天的close价
            if ticker in yesterday_close_cache:
                close_yesterday = yesterday_close_cache[ticker]
            else:
                close_yesterday = get_price_realtime(ticker, base_day)
                if close_yesterday is not None:
                    yesterday_close_cache[ticker] = close_yesterday
            if close_yesterday:
                change_7d = (close_yesterday - prev_close) / prev_close
                ws.cell(row=r, column=col_7d).value = round(change_7d, 4)
                ws.cell(row=r, column=col_7d).number_format = "0.00%"
                count_7d += 1
                print(f"7D 补齐: 行 {r}，股票 {ticker}，日期 {dt_val}，涨跌幅 {change_7d:.4%}")

    print(f"工作表 {sheet_name} 补齐结果：3D共 {count_3d} 条，7D共 {count_7d} 条")

print("补齐完成")
if "GITHUB_ACTIONS" in os.environ:
    os.system('rclone copy ./option_activity_log.xlsx "gdrive:/Investing/Daily top options" --drive-chunk-size 64M --progress --ignore-times')