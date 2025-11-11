import yfinance as yf
import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image as XLImage
import pytz
from datetime import datetime
from datetime import timedelta
import requests
import io

tz = pytz.timezone("America/Toronto")  # 先定义 tz
now = datetime.now(tz)                 # 然后才能用 tz
today = now.date()
yesterday = today - timedelta(days=1)
today_str = now.strftime('%Y-%m-%d')   # e.g., '2025-06-28'
now_time_str = now.strftime('%H:%M')   # e.g., '21:03'


# ==== 云端自动拉取最新 Excel ====
if "GITHUB_ACTIONS" in os.environ:
    os.system('rclone copy "gdrive:/Investing/Daily top options/option_activity_log.xlsx" ./ --drive-chunk-size 64M --progress --ignore-times')



print('📅 获取股票列表...')

# === 标普500 ===
url_sp500 = "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies"
headers = {"User-Agent": "Mozilla/5.0"} 
html_sp500 = requests.get(url_sp500, headers=headers).text
sp500_tables = pd.read_html(html_sp500)
sp500 = sp500_tables[0]

# 自动找包含 symbol/ticker 的列名
def find_symbol_column(df):
    for c in df.columns:
        if 'symbol' in c.lower() or 'ticker' in c.lower():
            return c
    raise KeyError(f"找不到 Symbol 或 Ticker 列，实际列名为: {list(df.columns)}")

sp500_symbol_col = find_symbol_column(sp500)
sp500_name_col = next((c for c in sp500.columns if 'security' in c.lower() or 'company' in c.lower()), None)

sp500_tickers = [t.replace('.', '-') for t in sp500[sp500_symbol_col].astype(str)]
sp500_names = dict(zip(sp500[sp500_symbol_col].astype(str).str.replace('.', '-'), sp500[sp500_name_col]))


# === 纳指100（自动识别正确表格） ===
url_nasdaq = "https://en.wikipedia.org/wiki/Nasdaq-100"
html_nasdaq = requests.get(url_nasdaq, headers=headers).text
nasdaq_tables = pd.read_html(html_nasdaq)

nasdaq100 = None
for df in nasdaq_tables:
    if any("ticker" in str(c).lower() or "symbol" in str(c).lower() for c in df.columns):
        nasdaq100 = df
        break
if nasdaq100 is None:
    raise ValueError(f"❌ 无法在 Nasdaq 页面中找到包含 ticker/symbol 的表格，实际表格数: {len(nasdaq_tables)}")

nasdaq_symbol_col = find_symbol_column(nasdaq100)
nasdaq_name_col = next((c for c in nasdaq100.columns if 'company' in c.lower()), None)

nasdaq100_tickers = [t.replace('.', '-') for t in nasdaq100[nasdaq_symbol_col].astype(str)]
nasdaq100_names = dict(zip(nasdaq100[nasdaq_symbol_col].astype(str).str.replace('.', '-'), nasdaq100[nasdaq_name_col]))



# === 纳斯达克全市场
url = "http://www.nasdaqtrader.com/dynamic/SymDir/nasdaqlisted.txt"
raw = requests.get(url).text
df_nasdaq = pd.read_csv(io.StringIO(raw), sep="|")[:-1]
df_nasdaq['Symbol'] = df_nasdaq['Symbol'].fillna("")
df_nasdaq = df_nasdaq[(df_nasdaq['ETF'] == 'N') & (~df_nasdaq['Symbol'].str.endswith(('U','W','R')))]
nasdaq_all_tickers = [t.replace('.', '-') for t in df_nasdaq['Symbol']]
nasdaq_all_names = dict(zip(df_nasdaq['Symbol'].str.replace('.', '-'), df_nasdaq['Security Name']))

# === 合并去重
all_tickers = set(sp500_tickers) | set(nasdaq100_tickers) | set(nasdaq_all_tickers)
all_tickers = sorted(list(all_tickers))  # 可选排序

# === ticker→公司名优先级（S&P500 > 纳指100 > 纳斯达克全市场）
ticker_name_map = {}
for t in all_tickers:
    if t in sp500_names:
        ticker_name_map[t] = sp500_names[t]
    elif t in nasdaq100_names:
        ticker_name_map[t] = nasdaq100_names[t]
    elif t in nasdaq_all_names:
        ticker_name_map[t] = nasdaq_all_names[t]
    else:
        ticker_name_map[t] = ""

tickers = all_tickers  # 你主程序直接用 tickers


records_raw = []
option_cache = {}

print("⏬ 批量拉取最近7天全部股票的历史收盘价 ...")
def batch_download(tickers, batch_size=500, **kwargs):
    all_df = []
    for i in range(0, len(tickers), batch_size):
        sub = tickers[i:i+batch_size]
        print(f"批量拉取 {i} ~ {i+len(sub)-1} ...")
        df = yf.download(sub, **kwargs)
        all_df.append(df)
    # 合并所有 batch
    df_all = pd.concat(all_df, axis=1)
    return df_all

price_df = batch_download(tickers, batch_size=500, period="7d", group_by="ticker")


# 模块顶层定义，循环外
def get_recent_close(stock, ref_date, max_lookback=7):
    for i in range(max_lookback):
        start_date = ref_date - timedelta(days=i+1)
        end_date = ref_date + timedelta(days=1)
        hist = stock.history(start=start_date.strftime('%Y-%m-%d'), end=end_date.strftime('%Y-%m-%d'))
        if not hist.empty:
            return hist['Close'].iloc[-1]
    return None

# 循环开始
for ticker in tickers:
    print(f'▶ 正在处理 {ticker} ...')
    try:
        stock = yf.Ticker(ticker)
        expiry_dates = stock.options

        # 获取昨天（或最近交易日）的收盘价
        close_price = get_recent_close(stock, yesterday)
        if close_price is not None:
            close_price = round(close_price, 2)

        # ✅ 计算涨跌幅（百分比）
        # ✅ 批量数据中计算2日收盘价（涨跌幅）
        try:
            closes = price_df[ticker]['Close'].dropna()
            if len(closes) >= 2:
                prev_close = closes.iloc[-2]
                last_price = closes.iloc[-1]
                price_change = round((last_price - prev_close) / prev_close, 4)
            else:
                price_change = None
        except Exception:
            price_change = None


# ✅ 计算 7 天价格变化（百分比）
        try:
            closes_7d = price_df[ticker]['Close'].dropna()
            if len(closes_7d) >= 2:
                week_start = closes_7d.iloc[0]
                week_end = closes_7d.iloc[-1]
                price_change_7d = round((week_end - week_start) / week_start, 4)
            else:
                price_change_7d = None
        except Exception:
            price_change_7d = None

        expiry_dates = [e for e in expiry_dates if (datetime.strptime(e, "%Y-%m-%d").date() - today).days <= 14]
        if not expiry_dates:
            continue

        all_calls, all_puts = [], []

        option_cache[ticker] = {}
        for expiry in expiry_dates:
            try:
        # 第一次才请求并缓存
                if expiry not in option_cache[ticker]:
                    chain = stock.option_chain(expiry)
                    option_cache[ticker][expiry] = {
                        'calls': chain.calls.copy(),
                        'puts': chain.puts.copy()
                    }
                calls = option_cache[ticker][expiry]['calls']
                puts = option_cache[ticker][expiry]['puts']
                calls['expiry'], puts['expiry'] = expiry, expiry
                all_calls.append(calls)
                all_puts.append(puts)
            except:
                continue


        if not all_calls or not all_puts:
            continue

        merged_calls = pd.concat(all_calls, ignore_index=True)
        merged_puts = pd.concat(all_puts, ignore_index=True)
        if merged_calls.empty or merged_puts.empty:
            continue

# ✅ 选出 top_call 和 top_put（成交量最大）
        top_call = merged_calls.sort_values('volume', ascending=False).iloc[0]
        top_put = merged_puts.sort_values('volume', ascending=False).iloc[0]

# ✅ IV Skew = Call IV - Put IV
        iv_skew = round(top_call['impliedVolatility'] * 100 - top_put['impliedVolatility'] * 100, 2)

        call_volume = top_call['volume']
        put_volume = top_put['volume']
        total_volume = merged_calls['volume'].sum() + merged_puts['volume'].sum()

        if total_volume < 3000:
            continue

        put_call_ratio = round(put_volume / call_volume, 4) if call_volume != 0 else float('inf')
        premium_skew = round(top_call['lastPrice'] - top_put['lastPrice'], 2)
        volume_diff_ratio = (call_volume - put_volume) / (call_volume + put_volume) if (call_volume + put_volume) != 0 else 0


# ✅ Premium Skew 打分（满分 36）
        if premium_skew >= 8:
            score_premium = 36
        elif premium_skew >= 5:
            score_premium = 30
        elif premium_skew >= 2:
            score_premium = 24
        elif premium_skew >= -1:
            score_premium = 18
        elif premium_skew >= -4:
            score_premium = 12
        elif premium_skew >= -7:
            score_premium = 6
        else:
            score_premium = 0

# ✅ IV Skew 打分（满分 30）
        if iv_skew >= 9:
            score_iv = 30
        elif iv_skew >= 5:
            score_iv = 25
        elif iv_skew >= 1:
            score_iv = 20
        elif iv_skew >= -3:
            score_iv = 15
        elif iv_skew >= -7:
            score_iv = 10
        elif iv_skew >= -11:
            score_iv = 5
        else:
            score_iv = 0

# ✅ Volume Diff Ratio 打分（满分 20）
        if volume_diff_ratio >= 0.7:
            score_vol = 20
        elif volume_diff_ratio >= 0.4:
            score_vol = 16
        elif volume_diff_ratio >= 0:
            score_vol = 12
        elif volume_diff_ratio >= -0.3:
            score_vol = 8
        elif volume_diff_ratio >= -0.6:
            score_vol = 4
        else:
            score_vol = 0

# ✅ Put/Call Ratio 打分（满分 14）
        if put_call_ratio <= 0.1:
            score_pcr = 14
        elif put_call_ratio <= 0.4:
            score_pcr = 12
        elif put_call_ratio <= 1:
            score_pcr = 10
        elif put_call_ratio <= 2:
            score_pcr = 8
        elif put_call_ratio <= 4:
            score_pcr = 6
        elif put_call_ratio <= 7:
            score_pcr = 4
        elif put_call_ratio <= 11:
            score_pcr = 2
        else:
            score_pcr = 0



        total_score = score_iv + score_premium + score_vol + score_pcr

        sentiment = (
            "Strong Bullish" if total_score >= 80 else
            "Bullish" if total_score >= 60 else
            "Neutral" if total_score >= 40 else
            "Bearish" if total_score >= 20 else
            "Strong Bearish"
        )

        for option_type, top_option in zip(['Call', 'Put'], [top_call, top_put]):
            records_raw.append({
                'Date': today_str,'Time': now_time_str,'Ticker': ticker, 'Company': ticker_name_map.get(ticker, ''),
                'Type': option_type, 'Strike': top_option['strike'],
                'IV': round(top_option['impliedVolatility'] * 100, 2),
                'Volume': int(top_option['volume']), 'Total Volume': total_volume, 
                'OI': int(top_option['openInterest']),
                'Expiry': top_option['expiry'],'Premium Skew': premium_skew,
                'IV Skew': iv_skew,'Volume Diff Ratio': round(volume_diff_ratio, 4), 
                'Put/Call Ratio': put_call_ratio, 'Score': total_score,
                'Sentiment': sentiment, 'Contract Symbol': top_option['contractSymbol'],
                'Previous Close': close_price, 'Price Change': price_change, '7D Change': price_change_7d})


    except Exception as e:
        continue

df = pd.DataFrame(records_raw)
print("收集到的数据行数：", len(df))
if df.empty:
    print("【警告】本次运行未获得任何有效期权数据，未写入Excel。")
    exit()

df = df[df['Total Volume'] > 3000]  # ✅ 集中过滤
top40 = df.groupby('Ticker')['Total Volume'].sum().nlargest(40).index
df = df[df['Ticker'].isin(top40)].copy()

df['TypeRank'] = df['Type'].apply(lambda x: 0 if x == 'Call' else 1)
df = df.sort_values(by=['Score', 'Ticker', 'TypeRank'], ascending=[False, True, True])
df.drop(columns=['Total Volume', 'TypeRank'], inplace=True)  # ✅ 排序后立即删掉，避免导出


sentiments = ["Strong Bullish", "Bullish", "Neutral", "Bearish", "Strong Bearish"]
sentiment_counts = df['Sentiment'].value_counts()

# ✅ 写入 Excel
file_name = "option_activity_log.xlsx"
month_sheet_name = now.strftime("%Y-%m")
year_sheet_name = now.strftime("%Y")

if not os.path.exists(file_name):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = month_sheet_name
    ws1.freeze_panes = 'D2'
    for r in dataframe_to_rows(df, index=False, header=True):
        ws1.append(r)
    # 不要加空行！！！（此时没有格式问题，月sheet可以直接写入）
    ws2 = wb.create_sheet(title=year_sheet_name)
    ws2.append(["Date", "Time", "Strong Bullish", "Bullish", "Neutral", "Bearish", "Strong Bearish", "Score"])
else:
    wb = load_workbook(file_name)
    if month_sheet_name in wb.sheetnames:
        ws1 = wb[month_sheet_name]
        last_data_row = ws1.max_row
    # 连续插入两行空白
        for _ in range(2):
            ws1.insert_rows(last_data_row + 1)
            last_data_row += 1
    # 定位写入开始行，两空行之后
        start_row = last_data_row + 1
        for i, row_data in enumerate(dataframe_to_rows(df, index=False, header=False)):
            for j, val in enumerate(row_data, start=1):
                ws1.cell(row=start_row + i, column=j, value=val)
    else:
        ws1 = wb.create_sheet(month_sheet_name)
        ws1.freeze_panes = 'D2'
        for r in dataframe_to_rows(df, index=False, header=True):
            ws1.append(r)
        # 不要加空行！！！

    if year_sheet_name in wb.sheetnames:
        ws2 = wb[year_sheet_name]
    else:
        ws2 = wb.create_sheet(year_sheet_name)
    ws2.freeze_panes = 'A2'


# ✅ 月Sheet上色
fills = {
    "Strong Bullish": "C6EFCE",
    "Bullish": "C6EFCE",
    "Neutral": "FFEB9C",
    "Bearish": "FFC7CE",
    "Strong Bearish": "FFC7CE"
}

sentiment_col_index = [c.value for c in ws1[1]].index("Sentiment") + 1
for row in ws1.iter_rows(min_row=2, min_col=sentiment_col_index, max_col=sentiment_col_index):
    for cell in row:
        cell.fill = PatternFill(start_color=fills.get(cell.value, "FFFFFF"), fill_type="solid")

# ✅ 设置百分比格式显示
for ws in [ws1]:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if ws.cell(row=1, column=cell.column).value in ['Price Change', '7D Change']:
                cell.number_format = '0.00%'  # ✅ 两位小数百分比格式

def get_last_data_row(ws, col=1):
    for row in range(ws.max_row, 0, -1):
        if ws.cell(row=row, column=col).value is not None:
            return row
    return 1  # 如果没数据，返回表头所在行

# 检查表头是否存在，不存在时添加
if ws2.max_row == 0:
    ws2.append(["Date", "Time", "Strong Bullish", "Bullish", "Neutral", "Bearish", "Strong Bearish", "Score"])

sentiments = ["Strong Bullish", "Bullish", "Neutral", "Bearish", "Strong Bearish"]
ticker_sentiment = df[['Ticker', 'Sentiment']].drop_duplicates(subset=['Ticker'])
sentiment_counts = ticker_sentiment['Sentiment'].value_counts()

sentiment_score = (
    sentiment_counts.get("Strong Bullish", 0) * 7 +
    sentiment_counts.get("Bullish", 0) * 5 +
    sentiment_counts.get("Neutral", 0) * 3 -
    sentiment_counts.get("Bearish", 0) * -
    sentiment_counts.get("Strong Bearish", 0) * (-3)
)

print(f"情绪统计：{sentiment_counts.to_dict()}")
print(f"计算综合分数：{sentiment_score}")

# 获取最后一行的行号
last_data_row = get_last_data_row(ws2, 1)
last_date = ws2.cell(row=last_data_row, column=1).value  # 第1列是 Date

print(f"[调试] last_data_row = {last_data_row}, last_date = {last_date}, now = {now.strftime('%Y-%m-%d')}")

if isinstance(last_date, str) and last_date != now.strftime("%Y-%m-%d"):
    for _ in range(2):
        ws2.insert_rows(last_data_row + 1)
        last_data_row += 1  # 每插入一行，更新 last_data_row，保证连续插入
    print("日期不同，追加两个空行作为分隔")

# 定位写入数据，替代原先的 append，确保数据写入空行后面
start_row = last_data_row + 1
data_to_write = [
    now.strftime("%Y-%m-%d"),
    now.strftime("%H:%M"),
    sentiment_counts.get("Strong Bullish", 0),
    sentiment_counts.get("Bullish", 0),
    sentiment_counts.get("Neutral", 0),
    sentiment_counts.get("Bearish", 0),
    sentiment_counts.get("Strong Bearish", 0),
    sentiment_score,
]

for col_idx, val in enumerate(data_to_write, start=1):
    ws2.cell(row=start_row, column=col_idx, value=val)

print(f"追加年sheet汇总行：{now.strftime('%Y-%m-%d %H:%M')}")


wb.save(file_name)

# 重新加载工作簿，确保数据已写入
wb = load_workbook(file_name)
ws2 = wb[year_sheet_name]

# ✅ 给第二页 Score 列添加红绿背景，表示分数变化（只对比最近两行）
score_col_index = [c.value for c in ws2[1]].index("Score") + 1
prev_score = None

for row in range(2, ws2.max_row + 1):  # 从第2行开始（表头是第1行）
    curr_score = ws2.cell(row=row, column=score_col_index).value
    if not isinstance(curr_score, int):
        continue  # 跳过空行
    if prev_score is not None:
        if curr_score > prev_score:
            ws2.cell(row=row, column=score_col_index).fill = PatternFill(start_color="C6EFCE", fill_type="solid")  # 绿
        elif curr_score < prev_score:
            ws2.cell(row=row, column=score_col_index).fill = PatternFill(start_color="FFC7CE", fill_type="solid")  # 红
    prev_score = curr_score  # 更新为当前分数，供下一行比较


wb.save(file_name)
print(f"✅ 数据已保存至 {file_name}")


# 读取文件画图
df_all = pd.read_excel(file_name, sheet_name=year_sheet_name)
df_all = df_all[df_all['Date'].notna()]
df_all['Date'] = pd.to_datetime(df_all['Date'])
df_all = df_all.sort_values('Date')

wb = load_workbook(file_name)
ws2 = wb[year_sheet_name]

ws2._images.clear()

max_col = ws2.max_column
start_col = max_col + 2

row_offset = 4
img_height_rows = 16

unique_dates = sorted(df_all['Date'].dt.date.unique())

for date_val in unique_dates:
    day_data = df_all[df_all['Date'].dt.date == date_val]
    day_data = day_data.sort_values('Time')
    x = day_data['Time']
    y = day_data['Score']

    plt.figure(figsize=(10,2))
    plt.plot(x, y, marker='o', linestyle='-')
    plt.title(f'Score on {date_val}')
    plt.xlabel('Time')
    plt.ylabel('Score')
    plt.xticks(x, x, rotation=45)
    plt.tight_layout()

    img_file = f'score_{date_val}.png'
    plt.savefig(img_file)
    plt.close()

    img = XLImage(img_file)
    img.anchor = f"{get_column_letter(start_col)}{1 + row_offset}"
    ws2.add_image(img)
    row_offset += img_height_rows



# ✅ 自动调整列宽
from openpyxl.utils import get_column_letter

for ws in [ws1, ws2]:
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        header = ws.cell(row=1, column=col[0].column).value
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[col_letter].width = max_length + 1

wb.save(file_name)  # 画图完毕后保存
print("✅ 数据和图形都已保存完成")

# ✅ 删除所有 score_*.png 文件（确保 save 之后才删）
png_files = [f for f in os.listdir() if f.startswith("score_") and f.endswith(".png")]
for f in png_files:
    try:
        os.remove(f)
    except Exception as e:
        print(f"⚠️ 无法删除 {f}: {e}")


if "GITHUB_ACTIONS" in os.environ:
    os.system('rclone copy ./option_activity_log.xlsx "gdrive:/Investing/Daily top options" --drive-chunk-size 64M --progress --ignore-times')