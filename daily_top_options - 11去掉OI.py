import yfinance as yf
import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

print('📅 获取股票列表...')
sp500 = pd.read_html('https://en.wikipedia.org/wiki/List_of_S%26P_500_companies')[0]
nasdaq = pd.read_html('https://en.wikipedia.org/wiki/Nasdaq-100')[4]
tickers = list(set(sp500['Symbol'].tolist() + nasdaq['Ticker'].tolist()))
tickers = [t.replace('.', '-') for t in tickers]

ticker_name_map = {}
for i in range(len(sp500)):
    symbol = sp500.loc[i, 'Symbol'].replace('.', '-')
    ticker_name_map[symbol] = sp500.loc[i, 'Security']
for i in range(len(nasdaq)):
    symbol = nasdaq.loc[i, 'Ticker'].replace('.', '-')
    if symbol not in ticker_name_map:
        ticker_name_map[symbol] = nasdaq.loc[i, 'Company']

records = []
volume_summary = {}
today_str = datetime.now().strftime('%Y-%m-%d')

for ticker in tickers:
    print(f'▶ 正在处理 {ticker} ...')
    try:
        stock = yf.Ticker(ticker)
        expiry_dates = stock.options
        today = datetime.today()
        expiry_dates = [e for e in expiry_dates if (datetime.strptime(e, "%Y-%m-%d") - today).days <= 10]
        if not expiry_dates:
            continue

        all_calls = []
        all_puts = []
        for expiry in expiry_dates:
            try:
                chain = stock.option_chain(expiry)
                calls = chain.calls.copy()
                puts = chain.puts.copy()
                calls['expiry'] = expiry
                puts['expiry'] = expiry
                all_calls.append(calls)
                all_puts.append(puts)
            except:
                continue

        if not all_calls or not all_puts:
            continue

        merged_calls = pd.concat(all_calls, ignore_index=True)
        merged_puts = pd.concat(all_puts, ignore_index=True)
        if merged_calls.empty or merged_puts.empty:
            continue

        top_call = merged_calls.sort_values('volume', ascending=False).iloc[0]
        top_put = merged_puts.sort_values('volume', ascending=False).iloc[0]

        call_volume = top_call['volume']
        put_volume = top_put['volume']
# ✅ 改成：不再用 top_call + top_put 来计算总成交量
        total_call_volume = merged_calls['volume'].sum()
        total_put_volume = merged_puts['volume'].sum()
        total_volume = total_call_volume + total_put_volume

# ✅ 原先是 continue，保留用 total_volume 来过滤垃圾 ticker
        if total_volume < 3000:
            continue

# ✅ 只在整支期权链完全没有 open interest 时才跳过continue
        put_call_ratio = round(put_volume / call_volume, 2) if call_volume != 0 else 0
        iv_skew = round(top_call['impliedVolatility'] * 100 - top_put['impliedVolatility'] * 100, 2)

# ✅ 情绪评分机制
        score = 0

# 1. Put/Call Ratio
        if put_call_ratio < 0.6:
            score += 1
        elif put_call_ratio > 1.4:
            score -= 1

# 2. 成交活跃度对比# 3. IV Skew 趋势
        if iv_skew > 2:
            score += 1
        elif iv_skew < -2:
            score -= 1

# 4. Call volume 显著大于 Put volume（放大倍数）
        if call_volume > 2 * put_volume:
            score += 1

# ✅ 最终情绪判断
        if score >= 2:
            sentiment = "Bullish"
        elif score <= -2:
            sentiment = "Bearish"
        else:
            sentiment = "Neutral"




        records.append({
            'Date': today_str, 'Ticker': ticker, 'Company': ticker_name_map.get(ticker, ''),
            'Type': 'Call', 'Strike': top_call['strike'], 'IV': round(top_call['impliedVolatility'] * 100, 2),
            'Volume': int(call_volume), 'Expiry': top_call['expiry'], 'Put/Call Ratio': put_call_ratio,
            'IV Skew': iv_skew, 'Sentiment': sentiment ,
            'contractSymbol': top_call['contractSymbol'] })

        records.append({
            'Date': today_str, 'Ticker': ticker, 'Company': ticker_name_map.get(ticker, ''),
            'Type': 'Put', 'Strike': top_put['strike'], 'IV': round(top_put['impliedVolatility'] * 100, 2),
            'Volume': int(put_volume), 'Expiry': top_put['expiry'], 'Put/Call Ratio': put_call_ratio,
            'IV Skew': iv_skew, 'Sentiment': sentiment ,
            'contractSymbol': top_put['contractSymbol']})

        volume_summary[ticker] = total_volume

    except Exception as e:
        print(f'⚠️ 跳过 {ticker}，错误：{e}')
        continue

df = pd.DataFrame(records)
if df.empty:
    print('❌ 最终 records 为空，没有生成有效数据')
    exit()

# ✅ 选出总成交量最多的前 40 个 ticker
top40_tickers = sorted(volume_summary.items(), key=lambda x: x[1], reverse=True)[:40]
top40_set = set(t[0] for t in top40_tickers)
df = df[df['Ticker'].isin(top40_set)]

# ✅ 排序逻辑：Ticker + Call优先 + 
df['TypeRank'] = df['Type'].map({'Call': 0, 'Put': 1})
df = df.sort_values(by=['Ticker', 'TypeRank'], ascending=[True, True])
df = df.drop(columns=['TypeRank'])

# ✅ 统计情绪
sentiment_counts = df['Sentiment'].value_counts()
bullish = sentiment_counts.get("Bullish", 0)
bearish = sentiment_counts.get("Bearish", 0)
neutral = sentiment_counts.get("Neutral", 0)
print(f"🧠 今日情绪分布：Bullish {bullish} / Bearish {bearish} / Neutral {neutral}")

# ✅ 写入 Excel
file_name = "option_activity_log.xlsx"
if not os.path.exists(file_name):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Top Options"
    ws1.freeze_panes = 'A2'  # ✅ 冻结第一行
    for r in dataframe_to_rows(df, index=False, header=True):
        ws1.append(r)
    ws1.append([])
    ws1.append([])
    ws2 = wb.create_sheet(title="Sentiment Stats")
    ws2.append(["Date", "Bullish", "Bearish", "Neutral"])
    ws2.append([today_str, bullish, bearish, neutral])
else:
    wb = load_workbook(file_name)
    ws1 = wb[wb.sheetnames[0]]
    ws1.freeze_panes = 'A2'  # ✅ 每次都冻结一次
    ws1.append([])
    ws1.append([])
    for r in dataframe_to_rows(df, index=False, header=False):
        ws1.append(r)
    if "Sentiment Stats" in wb.sheetnames:
        ws2 = wb["Sentiment Stats"]
    else:
        ws2 = wb.create_sheet(title="Sentiment Stats")
        ws2.append(["Date", "Bullish", "Bearish", "Neutral"])
    ws2.append([today_str, bullish, bearish, neutral])

# ✅ 着色
sentiment_col_index = None
for i, cell in enumerate(ws1[1], start=1):
    if cell.value == "Sentiment":
        sentiment_col_index = i
        break

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

if sentiment_col_index:
    for row in ws1.iter_rows(min_row=2, min_col=sentiment_col_index, max_col=sentiment_col_index):
        for cell in row:
            if cell.value == "Bullish":
                cell.fill = green_fill
            elif cell.value == "Bearish":
                cell.fill = red_fill
            elif cell.value == "Neutral":
                cell.fill = yellow_fill

# ✅ 自动调整列宽（新增）
from openpyxl.utils import get_column_letter

for col in ws1.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws1.column_dimensions[col_letter].width = max_length + 1


wb.save(file_name)
print(f"✅ 数据已保存至 {file_name}")