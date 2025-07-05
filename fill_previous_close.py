import pytz
from datetime import datetime, timedelta
import yfinance as yf
import os
from openpyxl import load_workbook

# ==== 云端自动拉取最新 Excel ====
if "GITHUB_ACTIONS" in os.environ:
    os.system('rclone copy "gdrive:/Investing/Daily top options/option_activity_log.xlsx" ./ --drive-chunk-size 64M --progress --ignore-times')


def get_previous_trading_close(stock, current_date, max_lookback=10):
    """
    给定某天，从前一天开始往前查最近交易日的收盘价，最多回溯 max_lookback 天
    """
    for i in range(1, max_lookback + 1):
        check_date = current_date - timedelta(days=i)
        start_str = check_date.strftime('%Y-%m-%d')
        end_str = (check_date + timedelta(days=1)).strftime('%Y-%m-%d')
        try:
            hist = stock.history(start=start_str, end=end_str)
            if not hist.empty:
                hist.index = hist.index.tz_localize(None)
                for dt in hist.index:
                    if dt.date() == check_date:
                        return float(hist.loc[dt, 'Close'])
        except:
            continue
    return None



tz = pytz.timezone("America/Toronto")
now = datetime.now(tz)
today_str = now.strftime("%Y-%m-%d")
file_name = "option_activity_log.xlsx"

wb = load_workbook(file_name)
sheet_names = wb.sheetnames

def is_month_sheet(name):
    try:
        datetime.strptime(name, "%Y-%m")
        return True
    except:
        return False

min_date = input("请输入补齐起始日期 (yyyy-mm-dd，留空不限制)：").strip()
max_date = input("请输入补齐截止日期 (yyyy-mm-dd，留空不限制)：").strip()

for sheet_name in sheet_names:
    if not is_month_sheet(sheet_name):
        continue

    ws = wb[sheet_name]
    header = [cell.value for cell in ws[1]]
    if "Previous Close" not in header or "Date" not in header or "Ticker" not in header:
        print(f"⚠️ Sheet {sheet_name} 缺少必要列，跳过")
        continue

    date_col = header.index("Date") + 1
    ticker_col = header.index("Ticker") + 1
    close_col = header.index("Previous Close") + 1

    fill_count = 0
    for r in range(2, ws.max_row + 1):
        date_cell = ws.cell(row=r, column=date_col).value
        ticker = ws.cell(row=r, column=ticker_col).value
        close_cell = ws.cell(row=r, column=close_col)

        if not ticker or not date_cell or str(close_cell.value).strip() not in ["", "None", "nan"]:
            continue

        row_date_str = str(date_cell)[:10]

        if min_date and row_date_str < min_date:
            continue
        if max_date and row_date_str > max_date:
            continue

        stock_symbol = str(ticker).strip()

        try:
            stock = yf.Ticker(stock_symbol)
            row_date = datetime.strptime(row_date_str, "%Y-%m-%d").date()
            prev_close = get_previous_trading_close(stock, row_date)
            if prev_close:
                print(f"✅ {stock_symbol}: got close {prev_close} (for {row_date_str})")
        except Exception as e:
            print(f"❌ {stock_symbol}: {e} (for {row_date_str})")



        if prev_close is not None:
            close_cell.value = round(prev_close, 2)
            close_cell.number_format = "0.00"  # 设置为两位小数
            fill_count += 1

    print(f"✅ Sheet {sheet_name}: 补齐 Previous Close 共 {fill_count} 条")

wb.save(file_name)
print("✅ 所有历史 Previous Close 均已补齐（今天除外，格式不破坏）")

# ==== 云端自动上传 Excel ====
if "GITHUB_ACTIONS" in os.environ:
    os.system('rclone copy ./option_activity_log.xlsx "gdrive:/Investing/Daily top options" --drive-chunk-size 64M --progress --ignore-times')
