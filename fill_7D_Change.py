import pandas as pd
from datetime import datetime, timedelta
import yfinance as yf
from openpyxl import load_workbook
import re

file_name = "option_activity_log.xlsx"
wb = load_workbook(file_name)

pattern = re.compile(r"^\d{4}-\d{2}$")  # 匹配 yyyy-mm 格式

def get_price_realtime(ticker, target_date, max_lookback=14):
    from datetime import timedelta
    for i in range(max_lookback):
        try_date = target_date - timedelta(days=i)
        if try_date.weekday() >= 5:  # 跳过周末
            continue
        try:
            stock = yf.Ticker(ticker)
            hist = stock.history(
                start=try_date.strftime("%Y-%m-%d"),
                end=(try_date + timedelta(days=1)).strftime("%Y-%m-%d"),
            )
            if not hist.empty:
                close_price = hist['Close'].iloc[0]
                print(f"获取 {ticker} 于 {try_date} 的收盘价: {close_price}")
                return close_price
        except Exception as e:
            print(f"错误: 获取 {ticker} 于 {try_date} 收盘价失败，原因: {e}")
            continue
    print(f"未找到 {ticker} 在 {target_date} 附近的有效收盘价")
    return None

print("开始补齐7D Change...")

for sheet_name in wb.sheetnames:
    if not pattern.match(sheet_name):
        continue
    print(f"处理工作表: {sheet_name}")
    ws = wb[sheet_name]
    header = [cell.value for cell in ws[1]]
    required_cols = ["Date", "Ticker", "Previous Close", "7D Change"]
    if any(col not in header for col in required_cols):
        print(f"工作表 {sheet_name} 缺少必要列，跳过")
        continue

    date_col = header.index("Date") + 1
    ticker_col = header.index("Ticker") + 1
    prev_close_col = header.index("Previous Close") + 1
    change_7d_col = header.index("7D Change") + 1

    count = 0
    cache_7d_price = {}

    for r in range(2, ws.max_row + 1):
        dt_cell = ws.cell(row=r, column=date_col).value
        if not dt_cell:
            continue
        if isinstance(dt_cell, datetime):
            dt_val = dt_cell.date()
        elif isinstance(dt_cell, str):
            dt_val = datetime.strptime(dt_cell[:10], "%Y-%m-%d").date()
        else:
            dt_val = dt_cell

        ticker = str(ws.cell(row=r, column=ticker_col).value).upper()
        prev_close = ws.cell(row=r, column=prev_close_col).value
        if prev_close is None or prev_close == 0:
            continue

        date_7d_ago = dt_val - timedelta(days=7)

        # 缓存7天前价格，避免多次请求
        cache_key = (ticker, date_7d_ago)
        if cache_key in cache_7d_price:
            price_7d_ago = cache_7d_price[cache_key]
        else:
            price_7d_ago = get_price_realtime(ticker, date_7d_ago)
            cache_7d_price[cache_key] = price_7d_ago

        if price_7d_ago:
            change_7d = (prev_close - price_7d_ago) / price_7d_ago
            ws.cell(row=r, column=change_7d_col).value = round(change_7d, 4)
            ws.cell(row=r, column=change_7d_col).number_format = "0.00%"
            count += 1
            print(f"行 {r} 股票 {ticker} 日期 {dt_val}: 7D Change={change_7d:.4%}")

    print(f"工作表 {sheet_name} 7D Change 补齐完成，共 {count} 条")

wb.save(file_name)
print("所有工作表7D Change补齐完成。")
