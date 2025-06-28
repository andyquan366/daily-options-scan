import yfinance as yf
import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

print('📅 获取股票列表...')
sp500 = pd.read_html('https://en.wikipedia.org/wiki/List_of_S%26P_500_companies')[0]
nasdaq = pd.read_html('https://en.wikipedia.org/wiki/Nasdaq-100')[4]
tickers = list(set(sp500['Symbol'].tolist() + nasdaq['Ticker'].tolist()))
tickers = [t.replace('.', '-') for t in tickers]

ticker_name_map = {}
for i in range(len(sp500)):
    symbol = sp500.loc[i, 'Symbol'].replace('.', '-')
    ticker_name_map[symbol] = sp500.loc[i, 'Security']
for i in range(len(nasdaq)):
    symbol = nasdaq.loc[i, 'Ticker'].replace('.', '-')
    if symbol not in ticker_name_map:
        ticker_name_map[symbol] = nasdaq.loc[i, 'Company']

records = []
volume_summary = {}
today_str = datetime.now().strftime('%Y-%m-%d')

for ticker in tickers:
    print(f'▶ 正在处理 {ticker} ...')
    try:
        stock = yf.Ticker(ticker)
        expiry_dates = stock.options
        today = datetime.today()
        expiry_dates = [e for e in expiry_dates if (datetime.strptime(e, "%Y-%m-%d") - today).days <= 7]
        if not expiry_dates:
            continue

        all_calls = []
        all_puts = []
        for expiry in expiry_dates:
            try:
                chain = stock.option_chain(expiry)
                calls = chain.calls.copy()
                puts = chain.puts.copy()
                calls['expiry'] = expiry
                puts['expiry'] = expiry
                all_calls.append(calls)
                all_puts.append(puts)
            except:
                continue

        if not all_calls or not all_puts:
            continue

        merged_calls = pd.concat(all_calls, ignore_index=True)
        merged_puts = pd.concat(all_puts, ignore_index=True)
        if merged_calls.empty or merged_puts.empty:
            continue

        top_call = merged_calls.sort_values('volume', ascending=False).iloc[0]
        top_put = merged_puts.sort_values('volume', ascending=False).iloc[0]

        call_volume = top_call['volume']
        call_oi = top_call['openInterest']
        put_volume = top_put['volume']
        put_oi = top_put['openInterest']

        if call_volume == 0 or put_volume == 0 or call_oi == 0 or put_oi == 0:
            continue

        put_call_ratio = round(put_volume / call_volume, 2)
        call_vo_ratio = round(call_volume / call_oi, 2)
        put_vo_ratio = round(put_volume / put_oi, 2)
        iv_skew = round(top_call['impliedVolatility'] * 100 - top_put['impliedVolatility'] * 100, 2)
        total_volume = call_volume + put_volume

        if put_call_ratio < 0.8 and call_vo_ratio > put_vo_ratio and iv_skew > 0:
            sentiment = 'Bullish'
        elif put_call_ratio > 1.2 and put_vo_ratio > call_vo_ratio and iv_skew < 0:
            sentiment = 'Bearish'
        else:
            sentiment = 'Neutral'

        volume_summary[ticker] = total_volume

        records.append({
            'Date': today_str, 'Ticker': ticker, 'Company': ticker_name_map.get(ticker, ''),
            'Type': 'Call', 'Strike': top_call['strike'], 'IV': round(top_call['impliedVolatility'] * 100, 2),
            'Volume': int(call_volume), 'Open Interest': int(call_oi), 'Volume/OI': call_vo_ratio,
            'Expiry': top_call['expiry'], 'Put/Call Ratio': put_call_ratio,
            'Call V/OI': call_vo_ratio, 'Put V/OI': put_vo_ratio, 'IV Skew': iv_skew, 'Sentiment': sentiment })

        records.append({
            'Date': today_str, 'Ticker': ticker, 'Company': ticker_name_map.get(ticker, ''),
            'Type': 'Put', 'Strike': top_put['strike'], 'IV': round(top_put['impliedVolatility'] * 100, 2),
            'Volume': int(put_volume), 'Open Interest': int(put_oi), 'Volume/OI': put_vo_ratio,
            'Expiry': top_put['expiry'], 'Put/Call Ratio': put_call_ratio,
            'Call V/OI': call_vo_ratio, 'Put V/OI': put_vo_ratio, 'IV Skew': iv_skew, 'Sentiment': sentiment })

    except Exception as e:
        print(f'⚠️ 跳过 {ticker}，错误：{e}')
        continue

df = pd.DataFrame(records)
if df.empty:
    print('❌ 最终 records 为空，没有生成有效数据')
    exit()

# ✅ 选出总成交量最多的前 20 个 ticker
top20_tickers = sorted(volume_summary.items(), key=lambda x: x[1], reverse=True)[:20]
top20_set = set(t[0] for t in top20_tickers)
df = df[df['Ticker'].isin(top20_set)]

# ✅ 排序逻辑：Ticker + Call优先 + Volume/OI降序
df['TypeRank'] = df['Type'].map({'Call': 0, 'Put': 1})
df = df.sort_values(by=['Ticker', 'TypeRank', 'Volume/OI'], ascending=[True, True, False])
df = df.drop(columns=['TypeRank'])

# ✅ 统计情绪
sentiment_counts = df['Sentiment'].value_counts()
bullish = sentiment_counts.get("Bullish", 0)
bearish = sentiment_counts.get("Bearish", 0)
neutral = sentiment_counts.get("Neutral", 0)
print(f"🧠 今日情绪分布：Bullish {bullish} / Bearish {bearish} / Neutral {neutral}")

# ✅ 写入 Excel
file_name = "option_activity_log.xlsx"
if not os.path.exists(file_name):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Top Options"
    for r in dataframe_to_rows(df, index=False, header=True):
        ws1.append(r)
    ws1.append([])
    ws1.append([])
    ws2 = wb.create_sheet(title="Sentiment Stats")
    ws2.append(["Date", "Bullish", "Bearish", "Neutral"])
    ws2.append([today_str, bullish, bearish, neutral])
else:
    wb = load_workbook(file_name)
    ws1 = wb[wb.sheetnames[0]]
    ws1.append([])
    ws1.append([])
    for r in dataframe_to_rows(df, index=False, header=False):
        ws1.append(r)
    if "Sentiment Stats" in wb.sheetnames:
        ws2 = wb["Sentiment Stats"]
    else:
        ws2 = wb.create_sheet(title="Sentiment Stats")
        ws2.append(["Date", "Bullish", "Bearish", "Neutral"])
    ws2.append([today_str, bullish, bearish, neutral])

# ✅ 着色
sentiment_col_index = None
for i, cell in enumerate(ws1[1], start=1):
    if cell.value == "Sentiment":
        sentiment_col_index = i
        break

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

if sentiment_col_index:
    for row in ws1.iter_rows(min_row=2, min_col=sentiment_col_index, max_col=sentiment_col_index):
        for cell in row:
            if cell.value == "Bullish":
                cell.fill = green_fill
            elif cell.value == "Bearish":
                cell.fill = red_fill
            elif cell.value == "Neutral":
                cell.fill = yellow_fill

wb.save(file_name)
print(f"✅ 数据已保存至 {file_name}")